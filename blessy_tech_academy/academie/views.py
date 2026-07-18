import hashlib
import json
import os
import random
from datetime import timedelta
from decimal import Decimal
from io import StringIO
import filetype
import markdown as markdown_lib
from django.conf import settings
from django.contrib import admin, messages
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core.management import call_command
from django.core.paginator import Paginator
from django.db import transaction as db_transaction
from django.db.models import Avg, Count, Q, Sum
from django.db.models.functions import TruncMonth
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.utils import timezone, translation
from django.views.decorators.cache import cache_page
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django_ratelimit.decorators import ratelimit
from drf_spectacular.utils import extend_schema, inline_serializer
from rest_framework import serializers
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from . import notifications
from .forms import ConnexionForm, InscriptionCompteForm, SujetForm
from .ia import (
    assistant_code,
    attribuer_badges,
    blessy_ai_repondre,
    chatbot_tuteur,
    correction_automatique,
    explication_concept,
    generateur_exercices,
    generer_contenu_formation,
    generer_contenu_lecon,
    generer_parcours_oriente,
    generer_programme_complet,
    generer_quiz,
    parcours_adaptatif,
    recommander_formations,
)
from .ia import simuler_carriere as simuler_carriere_ia
from .models import (
    AccesFormationDebloque,
    Article,
    Certificat,
    ChoixExamen,
    Coupon,
    Ecole,
    Enseignant,
    Examen,
    Formation,
    Inscription,
    InteractionCRM,
    Invoice,
    Lecon,
    Module,
    MoyenPaiement,
    Order,
    OrderItem,
    OutilRecommande,
    Parcours,
    ProgressionLecon,
    ProjetEtudiant,
    Promotion,
    Quiz,
    Question,
    Reaction,
    Refund,
    Reponse,
    ResultatQuiz,
    Sujet,
    Temoignage,
    TentativeExamen,
    Transaction,
    WorkflowFormation,
)
from .permissions import enregistrer_log, role_required
from .xp_utils import ajouter_xp


def _construire_contexte_utilisateur(request):
    """Construit un contexte utilisateur pour personnaliser les réponses du chatbot."""
    contexte_utilisateur = None
    if request.user.is_authenticated:
        formations_suivies = []
        progressions = (
            ProgressionLecon.objects.filter(utilisateur=request.user, terminee=True)
            .select_related("lecon__module__formation")
            .order_by("-date_completion")[:3]
        )
        for progression in progressions:
            formation = progression.lecon.module.formation
            if formation and formation.nom not in formations_suivies:
                formations_suivies.append(formation.nom)

        contexte_utilisateur = {
            "prenom": request.user.first_name or request.user.username,
            "formations_suivies": formations_suivies,
        }
    return contexte_utilisateur


# ================================================
# Pages principales
# ================================================


@cache_page(60 * 10)  # 10 minutes
def accueil(request):
    """Page d'accueil avec statistiques dynamiques."""
    formations = Formation.objects.filter(actif=True)[:4]
    formations_gratuites = Formation.objects.filter(actif=True, gratuit=True)
    ecoles = Ecole.objects.all()
    parcours_list = Parcours.objects.filter(actif=True)

    nb_etudiants = User.objects.filter(is_active=True).count()
    nb_formations = Formation.objects.filter(actif=True).count()
    nb_sujets_forum = Sujet.objects.count()

    stats = [
        {"valeur": nb_etudiants, "suffixe": "+", "label": "Étudiants"},
        {"valeur": nb_formations, "suffixe": "", "label": "Formations"},
        {"valeur": nb_sujets_forum, "suffixe": "", "label": "Sujets forum"},
    ]
    articles_recents = Article.objects.filter(publie=True).order_by("-date_publication")[:3]

    return render(
        request,
        "academie/accueil.html",
        {
            "formations": formations,
            "formations_gratuites": formations_gratuites,
            "ecoles": ecoles,
            "parcours_list": parcours_list,
            "stats": stats,
            "articles_recents": articles_recents,
            "nb_etudiants": nb_etudiants,
            "nb_formations": nb_formations,
            "nb_sujets_forum": nb_sujets_forum,
        },
    )


def apropos(request):
    return render(request, "academie/apropos.html")


def contact(request):
    from django import forms

    class ContactForm(forms.Form):
        prenom = forms.CharField(max_length=100, label="Prénom")
        nom = forms.CharField(max_length=100, label="Nom")
        email = forms.EmailField(label="Email")
        sujet = forms.CharField(max_length=200, label="Sujet")
        message = forms.CharField(widget=forms.Textarea, label="Message")

    if request.method == "POST":
        form = ContactForm(request.POST)
        if form.is_valid():
            messages.success(request, "✅ Message envoyé avec succès !")
            return redirect("contact")
    else:
        form = ContactForm()
    return render(request, "academie/contact.html", {"form": form})


@cache_page(60 * 15)  # 15 minutes
def formations(request):
    """Page des formations organisées par école + formations gratuites + parcours professionnels."""

    ecoles = Ecole.objects.prefetch_related("formations__modules").all()

    formations_gratuites = Formation.objects.filter(actif=True, gratuit=True)

    parcours_list = Parcours.objects.prefetch_related("formations").filter(actif=True)

    return render(
        request,
        "academie/formations.html",
        {
            "ecoles": ecoles,
            "formations_gratuites": formations_gratuites,
            "parcours_list": parcours_list,
        },
    )


# ================================================
# VIEWS.PY — detail_formation() enrichie pour le tunnel de conversion
# ================================================


def detail_formation(request, formation_id):
    """Page de détail d'une formation — tunnel de conversion premium."""
    formation = Formation.objects.prefetch_related("modules__lecons").get(
        id=formation_id, actif=True
    )

    pourcentage_progression = 0
    acces_autorise = formation.gratuit
    deja_inscrit = False

    if request.user.is_authenticated:
        pourcentage_progression = formation.progression_pour(request.user)
        acces_autorise = verifier_acces_formation(request.user, formation)
        deja_inscrit = acces_autorise

    # Prix dynamique avec promotion active
    prix_final, promo_active = _prix_avec_promotion(formation)

    # Statistiques du programme
    nb_modules = formation.modules.count()
    nb_lecons = sum(m.lecons.count() for m in formation.modules.all())
    duree_totale_minutes = sum(
        lecon.duree_minutes for m in formation.modules.all() for lecon in m.lecons.all()
    )

    # Formations similaires
    formations_similaires = (
        Formation.objects.filter(ecole=formation.ecole, actif=True).exclude(id=formation.id)[:3]
        if formation.ecole
        else []
    )

    # === Liste des débouchés (split par virgule ou point) ===
    if formation.debouches:
        formation.debouches_liste = [
            d.strip() for d in formation.debouches.replace(".", ",").split(",") if d.strip()
        ]
    else:
        formation.debouches_liste = []

    return render(
        request,
        "academie/detail_formation.html",
        {
            "formation": formation,
            "pourcentage_progression": pourcentage_progression,
            "acces_autorise": acces_autorise,
            "deja_inscrit": deja_inscrit,
            "prix_final": prix_final,
            "promo_active": promo_active,
            "nb_modules": nb_modules,
            "nb_lecons": nb_lecons,
            "duree_totale_heures": round(duree_totale_minutes / 60, 1),
            "formations_similaires": formations_similaires,
            "debouches_liste": formation.debouches_liste,
        },
    )


# ================================================
# VIEWS.PY — detail_formation_slug (SEO-friendly, réutilise detail_formation existante)
# Insérer juste après la fonction detail_formation() existante
# ================================================

def detail_formation_slug(request, formation_slug):
    """
    Version SEO-friendly de detail_formation — même logique,
    accessible via /formation/mon-slug-seo/ au lieu de /formation/42/.
    """
    formation = Formation.objects.filter(slug=formation_slug, actif=True).first()
    if not formation:
        from django.http import Http404
        raise Http404("Formation introuvable")

    # Délègue à la logique existante de detail_formation en interne
    return detail_formation(request, formation.id)


# ================================================
# Authentification
# ================================================


@ratelimit(key="ip", rate="3/m", block=True)
def inscription_compte(request):
    """Créer un nouveau compte étudiant."""
    if request.user.is_authenticated:
        return redirect("dashboard")

    if request.method == "POST":
        form = InscriptionCompteForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user, backend="django.contrib.auth.backends.ModelBackend")
            messages.success(
                request, f"🎉 Bienvenue {user.first_name} ! Ton compte a été créé avec succès."
            )
            return redirect("dashboard")
        else:
            messages.error(
                request,
                "❌ Erreur lors de la création du compte. Vérifie les informations saisies.",
            )
    else:
        form = InscriptionCompteForm()

    return render(request, "academie/inscription_compte.html", {"form": form})


@ratelimit(key="ip", rate="5/m", block=True)
def connexion(request):
    """Connexion à un compte existant."""
    if request.user.is_authenticated:
        return redirect("dashboard")

    if request.method == "POST":
        form = ConnexionForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user, backend="django.contrib.auth.backends.ModelBackend")
            messages.success(request, f"✅ Bienvenue {user.first_name or user.username} !")
            return redirect("dashboard")
        else:
            messages.error(
                request,
                "❌ Identifiants incorrects. Vérifie ton nom d'utilisateur et ton mot de passe.",
            )
    else:
        form = ConnexionForm(request)

    return render(request, "academie/connexion.html", {"form": form})


def deconnexion(request):
    """Déconnexion."""
    logout(request)
    messages.success(request, "👋 Tu as été déconnecté avec succès.")
    return redirect("accueil")


# ================================================
# VUE — Dashboard étudiant (badges, stats, examens)
# ================================================
@login_required(login_url="/connexion/")
def dashboard(request):
    """Tableau de bord étudiant moderne."""
    from .ia import attribuer_badges, calculer_stats_etudiant

    user = request.user
    stats = calculer_stats_etudiant(user)

    tous_badges = stats["badges"]

    formations_actives = sorted(stats["en_cours"], key=lambda f: f["pourcentage"], reverse=True)[:4]

    nouveaux_badges = attribuer_badges(user)
    if nouveaux_badges:
        messages.success(request, f"🎉 Nouveau(x) badge(s) : {', '.join(nouveaux_badges)} !")
    for badge_type in nouveaux_badges:
        notifications.notifier_badge(user, badge_type)

    from .models import ConnexionUtilisateur

    connexions = ConnexionUtilisateur.objects.filter(utilisateur=user).order_by("-date_connexion")[
        :5
    ]

    examens_passes = (
        TentativeExamen.objects.filter(utilisateur=user)
        .select_related("examen")
        .order_by("-date_debut")[:10]
    )

    return render(
        request,
        "academie/dashboard.html",
        {
            "user": user,
            "stats": stats,
            "badges": tous_badges,
            "formations_actives": formations_actives,
            "connexions": connexions,
            "examens_passes": examens_passes,
        },
    )


# ================================================
# VUE — Dashboard IA (resp_academique, direction, admin)
# ================================================
@login_required
@role_required("resp_academique", "direction", "admin", "super_admin")
def vue_dashboard_ia(request):
    from django.contrib.auth.models import User
    from django.db.models import Count, Q, Sum

    from .models import Article, Formation, Lecon, Order, ResultatQuiz

    formations_stats = (
        Formation.objects.filter(actif=True)
        .annotate(
            nb_inscrits=Count("orderitem", filter=Q(orderitem__commande__statut="paye")),
        )
        .order_by("-nb_inscrits")[:10]
    )

    lecons_abandon = Lecon.objects.annotate(
        nb_vues=Count("progressions"),
        nb_terminees=Count("progressions", filter=Q(progressions__terminee=True)),
    ).filter(nb_vues__gt=0)

    quiz_difficiles = (
        ResultatQuiz.objects.values("quiz__titre")
        .annotate(score_moyen=Avg("score"), nb_tentatives=Count("id"))
        .order_by("score_moyen")[:5]
    )

    articles_populaires = Article.objects.filter(publie=True).order_by("-date_publication")[:5]

    contexte_donnees = f"""
Formations les plus vendues : {[(f.nom, f.nb_inscrits) for f in formations_stats[:5]]}
Quiz avec scores les plus bas (difficiles) : {list(quiz_difficiles)}
Nombre total de formations actives : {Formation.objects.filter(actif=True).count()}
Nombre d'étudiants inscrits : {User.objects.filter(is_staff=False).count()}
Chiffre d'affaires : {Order.objects.filter(statut="paye").aggregate(t=Sum("total"))["t"] or 0} $
"""

    analyse_ia = None
    if request.GET.get("lancer_analyse"):
        from .ia import analyser_plateforme_ia

        analyse_ia = analyser_plateforme_ia(contexte_donnees)
    if analyse_ia:
        analyse_ia = analyse_ia.replace("##", "").replace("**", "").replace("---", "")
    return render(
        request,
        "admin/dashboard_ia.html",
        {
            "title": "🧠 Dashboard IA — Intelligence Décisionnelle",
            "formations_stats": formations_stats,
            "quiz_difficiles": quiz_difficiles,
            "articles_populaires": articles_populaires,
            "analyse_ia": analyse_ia,
        },
    )


# ================================================
# Statistiques (admin uniquement)
# ================================================


@login_required
@role_required("direction", "admin", "super_admin")
def statistiques(request):
    """Centre de Commande EdTech - Phases 1, 2 et 3."""
    from django.db.models import Sum

    # --- KPIs Phase 1 ---
    total_etudiants = User.objects.filter(is_active=True).count()
    total_formations = Formation.objects.filter(actif=True).count()
    total_certificats = Certificat.objects.count()
    total_leads = Inscription.objects.filter(formation__gratuit=True).count()
    total_inscriptions = Inscription.objects.count()
    taux_conversion = 0
    if total_leads > 0:
        taux_conversion = round((total_inscriptions / total_leads) * 100)

    # --- Tunnel Freemium (Phase 2) ---
    telechargements = total_leads
    comptes_crees = User.objects.count()
    inscriptions_payantes = Inscription.objects.filter(formation__gratuit=False).count()
    certificats_delivres = total_certificats
    taux_leads_comptes = (
        round((comptes_crees / telechargements) * 100) if telechargements > 0 else 0
    )
    taux_comptes_payant = (
        round((inscriptions_payantes / comptes_crees) * 100) if comptes_crees > 0 else 0
    )
    taux_payant_certif = (
        round((certificats_delivres / inscriptions_payantes) * 100)
        if inscriptions_payantes > 0
        else 0
    )

    # --- Revenus (Phase 2) ---
    total_revenus = (
        Inscription.objects.filter(formation__gratuit=False).aggregate(
            total=Sum("formation__prix")
        )["total"]
        or 0
    )

    # --- Certificats ce mois ---
    debut_mois = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    certificats_mois = Certificat.objects.filter(date_emission__gte=debut_mois).count()

    # --- IA Analytics (Phase 3) ---
    quiz_generees = Quiz.objects.count()
    contenus_generes = Lecon.objects.filter(contenu__isnull=False).exclude(contenu="").count()

    # --- Formations populaires ---
    formations_populaires = (
        Formation.objects.filter(actif=True)
        .annotate(nb_inscriptions=Count("inscriptions"))
        .order_by("-nb_inscriptions")[:5]
    )

    # --- Inscriptions par mois ---
    douze_mois = timezone.now() - timedelta(days=365)
    inscriptions_mensuelles = (
        Inscription.objects.filter(date_inscription__gte=douze_mois)
        .annotate(mois=TruncMonth("date_inscription"))
        .values("mois")
        .annotate(total=Count("id"))
        .order_by("mois")
    )
    mois_labels = [item["mois"].strftime("%b %Y") for item in inscriptions_mensuelles]
    mois_data = [item["total"] for item in inscriptions_mensuelles]

    # --- Activité forum ---
    total_sujets = Sujet.objects.count()
    total_reponses = Reponse.objects.count()
    membres_actifs = (
        User.objects.annotate(
            nb_actions=Count("sujets_forum", distinct=True) + Count("reponses_forum", distinct=True)
        )
        .filter(nb_actions__gt=0)
        .count()
    )

    # --- Centre d'alertes ---
    alertes = []
    inscriptions_non_traitees = Inscription.objects.filter(traite=False).count()
    if inscriptions_non_traitees > 0:
        alertes.append(f"{inscriptions_non_traitees} inscriptions non traitées")
    formations_sans_modules = Formation.objects.filter(actif=True, modules__isnull=True).count()
    if formations_sans_modules > 0:
        alertes.append(f"{formations_sans_modules} formations sans modules")
    etudiants_inactifs = User.objects.filter(is_active=True, progressions__isnull=True).count()
    if etudiants_inactifs > 0:
        alertes.append(f"{etudiants_inactifs} étudiants inactifs")

    contexte = {
        # Phase 1
        "total_etudiants": total_etudiants,
        "total_formations": total_formations,
        "total_certificats": total_certificats,
        "total_revenus": total_revenus,
        "total_leads": total_leads,
        "total_inscriptions": total_inscriptions,
        "taux_conversion": taux_conversion,
        "formations_populaires": formations_populaires,
        "mois_labels": mois_labels,
        "mois_data": mois_data,
        "total_sujets": total_sujets,
        "total_reponses": total_reponses,
        "membres_actifs": membres_actifs,
        "alertes": alertes,
        # Phase 2
        "telechargements": telechargements,
        "comptes_crees": comptes_crees,
        "inscriptions_payantes": inscriptions_payantes,
        "certificats_delivres": certificats_delivres,
        "taux_leads_comptes": taux_leads_comptes,
        "taux_comptes_payant": taux_comptes_payant,
        "taux_payant_certif": taux_payant_certif,
        "certificats_mois": certificats_mois,
        # Phase 3
        "quiz_generees": quiz_generees,
        "contenus_generes": contenus_generes,
    }
    return render(request, "academie/statistiques.html", contexte)


# ================================================
# Recherche
# ================================================


def recherche_formations(request):
    """Recherche de formations par mot-clé."""
    terme = request.GET.get("q", "")

    if terme:
        resultats = Formation.objects.filter(
            Q(nom__icontains=terme)
            | Q(description__icontains=terme)
            | Q(debouches__icontains=terme),
            actif=True,
        ).select_related("ecole")
    else:
        resultats = Formation.objects.none()

    return render(
        request,
        "academie/recherche.html",
        {
            "resultats": resultats,
            "terme": terme,
        },
    )


# ================================================
# Intelligence Artificielle — Blessy AI
# ================================================


def chat_ia(request):
    """Page du chat Blessy AI avec historique."""
    historique = request.session.get("chat_historique", [])
    return render(
        request,
        "academie/chat_ia.html",
        {
            "historique_chat": historique,
        },
    )


@ratelimit(key="user_or_ip", rate="10/m", block=True)
def api_chat_ia(request):
    """API endpoint pour le chat IA (AJAX) avec mémoire de conversation."""
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            question = data.get("question", "").strip()

            if not question:
                return JsonResponse({"erreur": "Question vide"}, status=400)

            if len(question) > 500:
                return JsonResponse(
                    {"erreur": "Question trop longue (max 500 caractères)"}, status=400
                )

            formations_actives = Formation.objects.filter(actif=True)[:5]
            historique = request.session.get("chat_historique", [])
            contexte_utilisateur = _construire_contexte_utilisateur(request)

            reponse = blessy_ai_repondre(
                question,
                formations_actives,
                historique=historique,
                contexte_utilisateur=contexte_utilisateur,
            )

            historique.append({"role": "user", "content": question})
            historique.append({"role": "assistant", "content": reponse})
            request.session["chat_historique"] = historique[-12:]
            request.session.modified = True

            # Convertir le Markdown en HTML
            try:
                reponse_html = markdown_lib.markdown(reponse)
            except Exception:
                reponse_html = reponse  # fallback si erreur

            return JsonResponse({"reponse": reponse_html})

        except Exception as e:
            return JsonResponse({"erreur": str(e)}, status=500)

    return JsonResponse({"erreur": "Méthode non autorisée"}, status=405)


def recommandations_ia(request):
    """Page de recommandations personnalisées."""
    recommandations = None
    interets = ""

    if request.method == "POST":
        interets = request.POST.get("interets", "").strip()
        if interets:
            formations_actives = Formation.objects.filter(actif=True)
            recommandations = recommander_formations(interets, formations_actives)

    return render(
        request,
        "academie/recommandations_ia.html",
        {
            "recommandations": recommandations,
            "interets": interets,
        },
    )


# ================================================
# API — Générer formation via IA (resp académique, admin)
# ================================================
@login_required
@role_required("resp_academique", "admin", "super_admin")
def api_generer_formation(request):
    """API pour générer le contenu d'une formation via l'IA."""
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            nom = data.get("nom", "").strip()
            ecole = data.get("ecole", "").strip()
            if not nom:
                return JsonResponse({"erreur": "Nom de formation requis"}, status=400)

            contenu = generer_contenu_formation(nom, ecole)

            return JsonResponse(contenu)

        except Exception as e:
            return JsonResponse({"erreur": str(e)}, status=500)

    return JsonResponse({"erreur": "Méthode non autorisée"}, status=405)


# ================================================
# API — Générer quiz via IA (resp académique, examinateur, admin)
# ================================================


@login_required
@role_required("resp_academique", "examinateur", "admin", "super_admin")
def api_generer_quiz(request):
    """API pour générer un quiz via l'IA."""
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            sujet = data.get("sujet", "").strip()
            nombre = int(data.get("nombre", 5))

            if not sujet:
                return JsonResponse({"erreur": "Sujet requis"}, status=400)

            questions = generer_quiz(sujet, nombre)

            if not questions:
                return JsonResponse({"erreur": "Génération échouée"}, status=500)

            return JsonResponse({"questions": questions})

        except Exception as e:
            return JsonResponse({"erreur": str(e)}, status=500)

    return JsonResponse({"erreur": "Méthode non autorisée"}, status=405)


# ================================================
# VIEWS.PY — Génération quiz au niveau Module (complète api_generer_quiz existante)
# ================================================
@login_required
@role_required('resp_academique', 'admin', 'super_admin')
def api_generer_quiz_module(request):
    """
    Génère un quiz rattaché à un MODULE spécifique (granularité fine).
    Réutilise generer_quiz() existante — juste le rattachement diffère.
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            module_id = data.get('module_id')
            nombre = int(data.get('nombre', 5))

            module = Module.objects.select_related('formation').get(id=module_id)

            questions = generer_quiz(module.titre, nombre)

            if not questions:
                return JsonResponse({'erreur': 'Génération échouée'}, status=500)

            quiz = Quiz.objects.create(
                formation=module.formation,
                module=module,
                titre=f"Quiz — {module.titre}",
                actif=True,
            )

            for i, q_data in enumerate(questions, start=1):
                Question.objects.create(
                    quiz=quiz,
                    texte=q_data.get('texte', ''),
                    choix_a=q_data.get('choix_a', ''),
                    choix_b=q_data.get('choix_b', ''),
                    choix_c=q_data.get('choix_c', ''),
                    choix_d=q_data.get('choix_d', ''),
                    bonne_reponse=q_data.get('bonne_reponse', 'a'),
                    explication=q_data.get('explication', ''),
                    ordre=i,
                )

            return JsonResponse({'succes': True, 'quiz_id': quiz.id, 'nombre_questions': len(questions)})

        except Exception as e:
            return JsonResponse({'erreur': str(e)}, status=500)

    return JsonResponse({'erreur': 'Méthode non autorisée'}, status=405)


def liste_quiz(request, formation_id):
    """Affiche les quiz disponibles pour une formation."""
    formation = Formation.objects.get(id=formation_id)
    quiz_disponibles = Quiz.objects.filter(formation=formation, actif=True)
    return render(
        request,
        "academie/liste_quiz.html",
        {
            "formation": formation,
            "quiz_disponibles": quiz_disponibles,
        },
    )


@login_required(login_url="/connexion/")
def passer_quiz(request, quiz_id):
    """Page pour passer un quiz."""
    quiz = Quiz.objects.prefetch_related("questions").get(id=quiz_id)

    if request.method == "POST":
        score = 0
        total = quiz.questions.count()

        for question in quiz.questions.all():
            reponse_utilisateur = request.POST.get(f"question_{question.id}")
            if reponse_utilisateur == question.bonne_reponse:
                score += 1

        ResultatQuiz.objects.create(
            utilisateur=request.user, quiz=quiz, score=score, total_questions=total
        )

        # Attribution des badges (toujours)
        attribuer_badges(request.user)

        # XP seulement si score >= 70%
        pourcentage = round((score / total) * 100) if total > 0 else 0
        if pourcentage >= 70:
            from .xp_utils import ajouter_xp

            ajouter_xp(request.user, "quiz_reussi")

            # Notification de succès
            notifications.creer_notification(
                request.user,
                "📝 Quiz réussi !",
                f'Tu as obtenu {score}/{total} au quiz "{quiz.titre}".',
                f"/formation/{quiz.formation.id}/quiz/",
            )

        return render(
            request,
            "academie/resultat_quiz.html",
            {
                "quiz": quiz,
                "score": score,
                "total": total,
                "pourcentage": pourcentage,
            },
        )

    return render(request, "academie/passer_quiz.html", {"quiz": quiz})


@login_required
@role_required("resp_academique", "admin", "super_admin")
def api_generer_programme(request):
    """API pour générer un programme complet (modules+leçons) via l'IA."""
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            formation_id = data.get("formation_id")

            if not formation_id:
                return JsonResponse({"erreur": "ID de formation requis"}, status=400)

            formation = Formation.objects.get(id=formation_id)

            programme = generer_programme_complet(
                formation.nom, formation.description, formation.niveau
            )

            if not programme:
                return JsonResponse({"erreur": "Génération échouée"}, status=500)

            for index_module, module_data in enumerate(programme, start=1):
                module = Module.objects.create(
                    formation=formation,
                    titre=module_data.get("titre", f"Module {index_module}"),
                    description=module_data.get("description", ""),
                    ordre=index_module,
                )

                for index_lecon, lecon_data in enumerate(module_data.get("lecons", []), start=1):
                    Lecon.objects.create(
                        module=module,
                        titre=lecon_data.get("titre", f"Leçon {index_lecon}"),
                        resume=lecon_data.get("resume", ""),
                        duree_minutes=lecon_data.get("duree_minutes", 15),
                        ordre=index_lecon,
                    )

            return JsonResponse(
                {
                    "succes": True,
                    "nombre_modules": len(programme),
                    "message": f"{len(programme)} modules créés avec succès !",
                }
            )

        except Formation.DoesNotExist:
            return JsonResponse({"erreur": "Formation introuvable"}, status=404)
        except Exception as e:
            return JsonResponse({"erreur": str(e)}, status=500)

    return JsonResponse({"erreur": "Méthode non autorisée"}, status=405)


# ================================================
# API — Générer contenu leçon via IA (formateur, resp académique, admin)
# ================================================
@login_required
@role_required("formateur", "resp_academique", "admin", "super_admin")
def api_generer_contenu_lecon(request):
    """API pour générer le contenu d'UNE leçon via l'IA."""
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            lecon_id = data.get("lecon_id")

            if not lecon_id:
                return JsonResponse({"erreur": "ID de leçon requis"}, status=400)

            lecon = Lecon.objects.select_related("module__formation").get(id=lecon_id)

            contenu = generer_contenu_lecon(
                titre_lecon=lecon.titre,
                resume_lecon=lecon.resume,
                contexte_formation=lecon.module.formation.nom,
                contexte_module=lecon.module.titre,
            )

            lecon.contenu = contenu
            lecon.save()

            return JsonResponse(
                {
                    "succes": True,
                    "contenu": contenu,
                }
            )

        except Lecon.DoesNotExist:
            return JsonResponse({"erreur": "Leçon introuvable"}, status=404)
        except Exception as e:
            return JsonResponse({"erreur": str(e)}, status=500)

    return JsonResponse({"erreur": "Méthode non autorisée"}, status=405)


@login_required
@role_required("resp_academique", "admin", "super_admin")
def api_generer_contenu_module(request):
    """API pour générer le contenu de TOUTES les leçons d'un module."""
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            module_id = data.get("module_id")

            if not module_id:
                return JsonResponse({"erreur": "ID de module requis"}, status=400)

            module = (
                Module.objects.select_related("formation")
                .prefetch_related("lecons")
                .get(id=module_id)
            )

            lecons = module.lecons.all()
            nombre_traitees = 0

            for lecon in lecons:
                contenu = generer_contenu_lecon(
                    titre_lecon=lecon.titre,
                    resume_lecon=lecon.resume,
                    contexte_formation=module.formation.nom,
                    contexte_module=module.titre,
                )
                lecon.contenu = contenu
                lecon.save()
                nombre_traitees += 1

            return JsonResponse(
                {
                    "succes": True,
                    "nombre_lecons": nombre_traitees,
                    "message": f"{nombre_traitees} leçons mises à jour avec succès !",
                }
            )

        except Module.DoesNotExist:
            return JsonResponse({"erreur": "Module introuvable"}, status=404)
        except Exception as e:
            return JsonResponse({"erreur": str(e)}, status=500)

    return JsonResponse({"erreur": "Méthode non autorisée"}, status=405)


# ================================================
# VIEWS.PY — lire_lecon() enrichie (préchargement + navigation clavier data)
# ================================================


@login_required(login_url="/connexion/")
def lire_lecon(request, lecon_id):
    """Page de lecture immersive — VERSION ENTERPRISE."""
    lecon = get_object_or_404(Lecon.objects.select_related("module__formation__ecole"), id=lecon_id)

    contenu_html = lecon.contenu if lecon.contenu else ""  # CKEditor génère déjà du HTML

    lecons_module = list(lecon.module.lecons.all())
    index_actuel = lecons_module.index(lecon)
    lecon_precedente = lecons_module[index_actuel - 1] if index_actuel > 0 else None
    lecon_suivante = (
        lecons_module[index_actuel + 1] if index_actuel < len(lecons_module) - 1 else None
    )

    progression = ProgressionLecon.objects.filter(utilisateur=request.user, lecon=lecon).first()
    lecon_terminee = progression.terminee if progression else False

    formation = lecon.module.formation
    pourcentage_formation = formation.progression_pour(request.user)

    # Table des matières complète (tous modules/leçons de la formation)
    tous_modules = formation.modules.prefetch_related("lecons").all()

    return render(
        request,
        "academie/lire_lecon.html",
        {
            "lecon": lecon,
            "contenu_html": contenu_html,
            "lecon_precedente": lecon_precedente,
            "lecon_suivante": lecon_suivante,
            "lecon_terminee": lecon_terminee,
            "pourcentage_formation": pourcentage_formation,
            "tous_modules": tous_modules,
            "formation": formation,
        },
    )


@login_required(login_url="/connexion/")
def marquer_lecon_terminee(request, lecon_id):
    """Marque une leçon comme terminée (ou non) pour l'utilisateur connecté."""
    if request.method == "POST":
        try:
            lecon = Lecon.objects.get(id=lecon_id)

            progression, cree = ProgressionLecon.objects.get_or_create(
                utilisateur=request.user,
                lecon=lecon,
            )

            progression.terminee = not progression.terminee
            progression.date_completion = timezone.now() if progression.terminee else None
            progression.save()

            # Ajouter XP seulement si la leçon est maintenant terminée
            if progression.terminee:
                ajouter_xp(request.user, "lecon_terminee")

                # Notification si la formation est maintenant complétée
                formation = lecon.module.formation
                pourcentage = formation.progression_pour(request.user)
                if pourcentage == 100:
                    notifications.notifier_formation_completee(request.user, formation.nom)

            formation = lecon.module.formation
            nouveau_pourcentage = formation.progression_pour(request.user)

            return JsonResponse(
                {
                    "succes": True,
                    "terminee": progression.terminee,
                    "progression_formation": nouveau_pourcentage,
                }
            )

        except Lecon.DoesNotExist:
            return JsonResponse({"erreur": "Leçon introuvable"}, status=404)
        except Exception as e:
            return JsonResponse({"erreur": str(e)}, status=500)

    return JsonResponse({"erreur": "Méthode non autorisée"}, status=405)


# ================================================
# Certificats PDF
# ================================================


@login_required(login_url="/connexion/")
def telecharger_certificat(request, formation_id):
    """Génère et télécharge le certificat PDF avec QR code de vérification."""
    import base64
    from io import BytesIO

    import qrcode

    from .models import Certificat

    formation = Formation.objects.prefetch_related("modules__lecons").get(
        id=formation_id, actif=True
    )

    pourcentage = formation.progression_pour(request.user)

    if pourcentage < 100:
        messages.error(
            request,
            f"❌ Tu dois compléter 100% de la formation pour obtenir le certificat "
            f"(progression actuelle : {pourcentage}%).",
        )
        return redirect("detail_formation", formation_id=formation_id)

    # Génère un numéro de certificat unique
    chaine = f"{request.user.id}-{formation.id}-{request.user.date_joined}"
    numero = f"BTA-{hashlib.md5(chaine.encode()).hexdigest()[:8].upper()}"

    # Crée le certificat en base s'il n'existe pas
    certificat, created = Certificat.objects.get_or_create(
        utilisateur=request.user, formation=formation, defaults={"numero": numero}
    )
    # Si déjà existant, utiliser le numéro existant
    if not created:
        numero = certificat.numero

    # URL de vérification
    url_verification = request.build_absolute_uri(f"/certificat/{numero}/")

    # Génère le QR code
    qr = qrcode.QRCode(version=1, box_size=10, border=2)
    qr.add_data(url_verification)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")

    # Convertit l'image en base64 pour l'insérer dans le HTML
    buffer = BytesIO()
    img.save(buffer, format="PNG")
    qr_code_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")

    contexte = {
        "prenom": request.user.first_name or request.user.username,
        "nom": request.user.last_name or "",
        "formation": formation,
        "date_emission": certificat.date_emission.strftime("%d %B %Y"),
        "numero_certificat": numero,
        "qr_code_base64": qr_code_base64,
        "url_verification": url_verification,
    }

    html_certificat = render_to_string("academie/pdf/certificat.html", contexte, request=request)

    try:
        from weasyprint import HTML

        pdf = HTML(string=html_certificat, base_url=request.build_absolute_uri("/")).write_pdf()

        nom_fichier = f"certificat-{formation.nom.replace(' ', '-').lower()}-BTA.pdf"
        response = HttpResponse(pdf, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{nom_fichier}"'
        return response

    except Exception as e:
        messages.error(request, f"❌ Erreur lors de la génération du certificat : {str(e)}")
        return redirect("detail_formation", formation_id=formation_id)

    # ================================================


# Forum Communautaire
# ================================================


# VUE — Forum (liste des sujets, optimisé N+1)
def forum_liste(request):
    """Page principale du forum — liste tous les sujets."""
    from django.db.models import Count as DjangoCount

    categorie = request.GET.get("categorie", "")
    formation_id = request.GET.get("formation", "")
    recherche = request.GET.get("q", "")

    sujets = Sujet.objects.select_related("auteur", "formation").annotate(
        nb_reponses_calc=DjangoCount("reponses", distinct=True),
        nb_likes_calc=DjangoCount("reactions", distinct=True),
    )

    if categorie:
        sujets = sujets.filter(categorie=categorie)

    if formation_id:
        sujets = sujets.filter(formation_id=formation_id)

    if recherche:
        sujets = sujets.filter(Q(titre__icontains=recherche) | Q(contenu__icontains=recherche))

    paginator = Paginator(sujets, 10)
    page_number = request.GET.get("page", 1)
    page_obj = paginator.get_page(page_number)

    formations = Formation.objects.filter(actif=True)

    return render(
        request,
        "academie/forum/liste.html",
        {
            "sujets": page_obj,
            "page_obj": page_obj,
            "formations": formations,
            "categorie_active": categorie,
            "formation_active": formation_id,
            "categories": Sujet.CATEGORIES,
            "recherche": recherche,
        },
    )


def forum_detail(request, sujet_id):
    """Page de détail d'un sujet avec ses réponses."""
    sujet = (
        Sujet.objects.select_related("auteur", "formation")
        .prefetch_related("reponses__auteur", "reponses__reactions", "reactions")
        .get(id=sujet_id)
    )

    # Incrémente le compteur de vues
    sujet.vues += 1
    sujet.save(update_fields=["vues"])

    # Likes de l'utilisateur connecté
    likes_sujets = set()
    likes_reponses = set()

    if request.user.is_authenticated:
        likes_sujets = set(
            Reaction.objects.filter(utilisateur=request.user, sujet=sujet).values_list(
                "sujet_id", flat=True
            )
        )
        likes_reponses = set(
            Reaction.objects.filter(utilisateur=request.user, reponse__sujet=sujet).values_list(
                "reponse_id", flat=True
            )
        )

    # Traitement du formulaire de réponse
    if request.method == "POST" and request.user.is_authenticated:
        contenu = request.POST.get("contenu", "").strip()
        if contenu:
            Reponse.objects.create(
                sujet=sujet,
                contenu=contenu,
                auteur=request.user,
            )

            # Ajouter XP pour la réponse postée
            ajouter_xp(request.user, "reponse_forum")

            messages.success(request, "✅ Réponse publiée !")
            return redirect("forum_detail", sujet_id=sujet_id)

    return render(
        request,
        "academie/forum/detail.html",
        {
            "sujet": sujet,
            "likes_sujets": likes_sujets,
            "likes_reponses": likes_reponses,
        },
    )


@login_required(login_url="/connexion/")
def forum_liker(request, type_cible, cible_id):
    """Toggle like sur un sujet ou une réponse."""
    if request.method == "POST":
        try:
            if type_cible == "sujet":
                sujet = Sujet.objects.get(id=cible_id)
                reaction, cree = Reaction.objects.get_or_create(
                    utilisateur=request.user,
                    sujet=sujet,
                )
                if not cree:
                    reaction.delete()
                    liked = False
                else:
                    liked = True
                total = sujet.reactions.count()

            elif type_cible == "reponse":
                reponse = Reponse.objects.get(id=cible_id)
                reaction, cree = Reaction.objects.get_or_create(
                    utilisateur=request.user,
                    reponse=reponse,
                )
                if not cree:
                    reaction.delete()
                    liked = False
                else:
                    liked = True
                total = reponse.reactions.count()

            else:
                return JsonResponse({"erreur": "Type invalide"}, status=400)

            return JsonResponse(
                {
                    "succes": True,
                    "liked": liked,
                    "total": total,
                }
            )

        except Exception as e:
            return JsonResponse({"erreur": str(e)}, status=500)

    return JsonResponse({"erreur": "Méthode non autorisée"}, status=405)


@login_required(login_url="/connexion/")
def forum_accepter_reponse(request, reponse_id):
    """Marque une réponse comme solution acceptée."""
    if request.method == "POST":
        reponse = Reponse.objects.select_related("sujet").get(id=reponse_id)

        # Seul l'auteur du sujet peut accepter une réponse
        if request.user != reponse.sujet.auteur:
            messages.error(request, "❌ Seul l'auteur du sujet peut accepter une réponse.")
            return redirect("forum_detail", sujet_id=reponse.sujet.id)

        # Désactive toutes les autres réponses acceptées
        Reponse.objects.filter(sujet=reponse.sujet).update(acceptee=False)

        # Active celle-ci
        reponse.acceptee = True
        reponse.save()

        # Marque le sujet comme résolu
        reponse.sujet.resolu = True
        reponse.sujet.save()
        attribuer_badges(reponse.auteur)

        # Ajouter XP pour la réponse acceptée
        ajouter_xp(reponse.auteur, "reponse_acceptee")

        notifications.creer_notification(
            reponse.auteur,
            "✅ Réponse acceptée",
            f'Ta réponse sur "{reponse.sujet.titre}" a été acceptée comme solution.',
            f"/forum/{reponse.sujet.id}/",
        )
        messages.success(request, "✅ Réponse marquée comme solution !")
        return redirect("forum_detail", sujet_id=reponse.sujet.id)

    return redirect("forum_liste")


@login_required(login_url="/connexion/")
def forum_creer(request):
    """Créer un nouveau sujet."""
    if request.method == "POST":
        titre = request.POST.get("titre", "").strip()
        contenu = request.POST.get("contenu", "").strip()
        categorie = request.POST.get("categorie", "general")
        formation_id = request.POST.get("formation", "")

        if not titre or not contenu:
            messages.error(request, "❌ Titre et contenu sont obligatoires.")
            formations = Formation.objects.filter(actif=True)
            return render(
                request,
                "academie/forum/creer.html",
                {
                    "form": SujetForm(),
                    "formations": formations,
                    "categories": Sujet.CATEGORIES,
                },
            )

        sujet = Sujet.objects.create(
            titre=titre,
            contenu=contenu,
            categorie=categorie,
            auteur=request.user,
            formation_id=formation_id if formation_id else None,
        )

        attribuer_badges(request.user)
        notifications.creer_notification(
            request.user,
            "📝 Sujet créé",
            f'Ton sujet "{sujet.titre}" a été publié avec succès.',
            f"/forum/{sujet.id}/",
        )
        messages.success(request, "✅ Sujet créé avec succès !")
        return redirect("forum_detail", sujet_id=sujet.id)

    formations = Formation.objects.filter(actif=True)
    return render(
        request,
        "academie/forum/creer.html",
        {
            "form": SujetForm(),
            "formations": formations,
            "categories": Sujet.CATEGORIES,
        },
    )


def forum_membres(request):
    """Classement des membres du forum."""
    from django.contrib.auth.models import User

    membres = (
        User.objects.annotate(
            nb_sujets=Count("sujets_forum", distinct=True),
            nb_reponses=Count("reponses_forum", distinct=True),
            nb_solutions=Count(
                "reponses_forum", filter=Q(reponses_forum__acceptee=True), distinct=True
            ),
            nb_likes=Count("reactions_forum", distinct=True),
        )
        .filter(Q(nb_sujets__gt=0) | Q(nb_reponses__gt=0))
        .order_by("-nb_reponses", "-nb_solutions", "-nb_sujets")[:20]
    )

    return render(
        request,
        "academie/forum/membres.html",
        {
            "membres": membres,
        },
    )


# ================================================
# Parcours d'Orientation Intelligent
# ================================================


def orientation_ia(request):
    """Page du parcours d'orientation personnalisé."""

    profils = [
        ("lyceen_etudiant", "Lycéen / Étudiant", "🎓"),
        ("professionnel", "Professionnel", "💼"),
        ("entrepreneur", "Entrepreneur", "🚀"),
        ("numerique", "Déjà dans le numérique", "👨‍💻"),
    ]

    objectifs = [
        ("developpeur", "Devenir développeur", "💻"),
        ("design_creation", "Design & Création", "🎨"),
        ("marketing_business", "Marketing & Business", "📊"),
        ("maitriser_ia", "Maîtriser l'IA", "🤖"),
        ("technicien", "Technicien informatique", "🔧"),
    ]

    disponibilites = [
        ("1-2h", "1-2h par jour", "⏰"),
        ("3-4h", "3-4h par jour", "⏱"),
        ("temps_plein", "Temps plein", "🔥"),
    ]

    resultat = None
    erreur = None

    if request.method == "POST":
        profil = request.POST.get("profil", "")
        objectif = request.POST.get("objectif", "")
        disponibilite = request.POST.get("disponibilite", "")
        details = request.POST.get("details", "").strip()

        if profil and objectif and disponibilite:
            formations = Formation.objects.filter(actif=True).select_related("ecole")
            resultat = generer_parcours_oriente(
                profil, objectif, disponibilite, details, formations
            )

            if "erreur" in resultat:
                erreur = resultat["erreur"]
                resultat = None
        else:
            erreur = "Veuillez répondre aux 3 premières questions."

    return render(
        request,
        "academie/orientation.html",
        {
            "resultat": resultat,
            "erreur": erreur,
            "profils": profils,
            "objectifs": objectifs,
            "disponibilites": disponibilites,
            "post_data": request.POST if request.method == "POST" else None,
        },
    )


# ================================================
# VUES API POUR LE BOUTON IA (6 fonctionnalités)
# ================================================


@require_POST
def api_assistant_code(request):
    """API 1 : Assistant Code - Analyse et corrige le code"""
    try:
        data = json.loads(request.body)
        code = data.get("code", "")
        langage = data.get("langage", "python")
        question = data.get("question", "")

        if not code:
            return JsonResponse({"erreur": 'Le champ "code" est requis.'}, status=400)

        reponse = assistant_code(code=code, langage=langage, question=question)
        return JsonResponse({"reponse": reponse, "fonction": "assistant_code"})
    except json.JSONDecodeError:
        return JsonResponse({"erreur": "JSON invalide"}, status=400)
    except Exception as e:
        return JsonResponse({"erreur": str(e)}, status=500)


@require_POST
def api_generateur_exercices(request):
    """API 2 : Générateur d'exercices personnalisés"""
    try:
        data = json.loads(request.body)
        sujet = data.get("sujet", "")
        niveau = data.get("niveau", "debutant")
        format_ex = data.get("format_exercice", "code")

        if not sujet:
            return JsonResponse({"erreur": 'Le champ "sujet" est requis.'}, status=400)

        reponse = generateur_exercices(sujet=sujet, niveau=niveau, format_exercice=format_ex)
        return JsonResponse({"reponse": reponse, "fonction": "generateur_exercices"})
    except json.JSONDecodeError:
        return JsonResponse({"erreur": "JSON invalide"}, status=400)
    except Exception as e:
        return JsonResponse({"erreur": str(e)}, status=500)


@require_POST
def api_explication_concept(request):
    """API 3 : Explication de concept"""
    try:
        data = json.loads(request.body)
        question = data.get("question", "")
        niveau_eleve = data.get("niveau_eleve", "debutant")

        if not question:
            return JsonResponse({"erreur": 'Le champ "question" est requis.'}, status=400)

        reponse = explication_concept(question=question, niveau_eleve=niveau_eleve)
        return JsonResponse({"reponse": reponse, "fonction": "explication_concept"})
    except json.JSONDecodeError:
        return JsonResponse({"erreur": "JSON invalide"}, status=400)
    except Exception as e:
        return JsonResponse({"erreur": str(e)}, status=500)


@require_POST
def api_correction_automatique(request):
    """API 4 : Correction automatique d'exercice"""
    try:
        data = json.loads(request.body)
        enonce = data.get("enonce", "")
        reponse_eleve = data.get("reponse_eleve", "")
        bareme = data.get("bareme", "")

        if not enonce or not reponse_eleve:
            return JsonResponse(
                {"erreur": 'Les champs "enonce" et "reponse_eleve" sont requis.'}, status=400
            )

        reponse = correction_automatique(enonce=enonce, reponse_eleve=reponse_eleve, bareme=bareme)
        return JsonResponse({"reponse": reponse, "fonction": "correction_automatique"})
    except json.JSONDecodeError:
        return JsonResponse({"erreur": "JSON invalide"}, status=400)
    except Exception as e:
        return JsonResponse({"erreur": str(e)}, status=500)


@require_POST
def api_parcours_adaptatif(request):
    """API 5 : Recommandation de parcours adaptatif"""
    try:
        data = json.loads(request.body)
        profil_scores = data.get("profil_scores", {})
        parcours_actuel = data.get("parcours_actuel", "")
        objectif = data.get("objectif", "")

        if not profil_scores:
            return JsonResponse({"erreur": 'Le champ "profil_scores" est requis.'}, status=400)

        reponse = parcours_adaptatif(
            profil_scores=profil_scores, parcours_actuel=parcours_actuel, objectif=objectif
        )
        return JsonResponse({"reponse": reponse, "fonction": "parcours_adaptatif"})
    except json.JSONDecodeError:
        return JsonResponse({"erreur": "JSON invalide"}, status=400)
    except Exception as e:
        return JsonResponse({"erreur": str(e)}, status=500)


@require_POST
def api_chatbot_tuteur(request):
    """API 6 : Chatbot tuteur conversationnel"""
    try:
        data = json.loads(request.body)
        message = data.get("message", "")
        historique = data.get("historique", [])
        niveau_eleve = data.get("niveau_eleve", "debutant")

        if not message:
            return JsonResponse({"erreur": 'Le champ "message" est requis.'}, status=400)

        reponse = chatbot_tuteur(message=message, historique=historique, niveau_eleve=niveau_eleve)
        return JsonResponse({"reponse": reponse, "fonction": "chatbot_tuteur"})
    except json.JSONDecodeError:
        return JsonResponse({"erreur": "JSON invalide"}, status=400)
    except Exception as e:
        return JsonResponse({"erreur": str(e)}, status=500)


# ================================================
# Simulateur de carrière
# ================================================


def simuler_carriere(request):
    """Simulateur de carrière — orientation IA ultra-détaillée."""

    METIERS_BTA = {
        "developpeur_web": {
            "titre": "Développeur Web Full Stack",
            "emoji": "💻",
            "description": "Tu crées des sites et applications web de A à Z — du design à la base de données.",
            "competences": ["HTML/CSS", "JavaScript", "Python", "Django", "PostgreSQL", "Git"],
            "technologies": ["VS Code", "GitHub", "Figma", "Postman", "Docker"],
            "metiers_accessibles": [
                "Développeur Front-end",
                "Développeur Back-end",
                "Full Stack Developer",
                "CTO Startup",
                "Freelance",
                "Consultant Web",
            ],
            "salaire_haiti": "800-2500",
            "salaire_international": "3000-8000",
            "duree_formation": "10 mois",
            "parcours_bta": "Développeur Web Python",
            "secteurs": ["Startups Tech", "Agences Web", "ONG", "Banques", "Télétravail"],
            "perspectives": "Évolution vers Lead Developer, Architecte logiciel, CTO ou fondateur de startup.",
            "demande": "Le développement web est l'une des compétences les plus recherchées au monde avec +25% de croissance annuelle.",
            "freelance": "Excellent — plateforme Upwork, Fiverr, Toptal. Revenus freelance dès le 6e mois.",
            "competences_futur": ["IA générative", "Web3", "Cloud Computing", "DevOps"],
        },
        "expert_ia": {
            "titre": "Expert en Intelligence Artificielle",
            "emoji": "🤖",
            "description": "Tu maîtrises les outils IA les plus puissants et tu les appliques dans des contextes professionnels concrets.",
            "competences": [
                "Prompt Engineering",
                "ChatGPT",
                "Claude",
                "Gemini",
                "Python IA",
                "Data Analysis",
            ],
            "technologies": ["OpenAI API", "LangChain", "Hugging Face", "Google Colab"],
            "metiers_accessibles": [
                "Prompt Engineer",
                "AI Product Manager",
                "Data Analyst",
                "Consultant IA",
                "Formateur IA",
            ],
            "salaire_haiti": "1500-4000",
            "salaire_international": "4000-12000",
            "duree_formation": "6 mois",
            "parcours_bta": "Spécialiste IA et Productivité",
            "secteurs": ["Finance", "Santé", "Éducation", "Marketing", "Consulting", "Remote"],
            "perspectives": "Secteur en explosion — les experts IA sont parmi les professionnels les mieux payés au monde.",
            "demande": "La demande en expertise IA croît de +40% par an. Les entreprises cherchent désespérément ces profils.",
            "freelance": "Excellent — services de consultation IA très demandés. Tarifs: 50-200 USD/heure.",
            "competences_futur": ["AGI", "IA multimodale", "IA embarquée", "Éthique IA"],
        },
        "technicien_informatique": {
            "titre": "Technicien Informatique",
            "emoji": "🖥️",
            "description": "Tu répares, configures et maintiens les équipements informatiques et réseaux des entreprises.",
            "competences": [
                "Hardware",
                "Windows Server",
                "Réseaux",
                "Dépannage",
                "Sécurité de base",
            ],
            "technologies": ["Active Directory", "VMware", "Cisco", "TeamViewer", "SCCM"],
            "metiers_accessibles": [
                "Technicien IT",
                "Support N1/N2",
                "Admin Réseau Junior",
                "Technicien Télécoms",
            ],
            "salaire_haiti": "600-1800",
            "salaire_international": "2500-5000",
            "duree_formation": "9 mois",
            "parcours_bta": "Technicien Informatique Professionnel",
            "secteurs": ["PME", "Hôtels", "Banques", "Hôpitaux", "Écoles", "ONG"],
            "perspectives": "Évolution vers Admin Systèmes, Ingénieur Réseau, Cybersécurité.",
            "demande": "Besoin constant dans toutes les organisations. Pénurie de techniciens qualifiés en Haïti.",
            "freelance": "Possible — maintenance informatique à domicile, support IT aux PME.",
            "competences_futur": ["Cloud hybride", "IoT", "Cybersécurité", "IA pour IT"],
        },
        "designer_graphique": {
            "titre": "Designer Graphique & Content Creator",
            "emoji": "🎨",
            "description": "Tu crées des identités visuelles, des contenus pour les réseaux sociaux et du matériel marketing professionnel.",
            "competences": [
                "Canva Pro",
                "Adobe Express",
                "Photoshop",
                "Illustrator",
                "Branding",
                "Vidéo",
            ],
            "technologies": ["Adobe Suite", "Figma", "CapCut", "DaVinci Resolve", "Midjourney"],
            "metiers_accessibles": [
                "Graphiste Freelance",
                "Community Manager",
                "Social Media Manager",
                "Brand Designer",
            ],
            "salaire_haiti": "500-1500",
            "salaire_international": "2000-6000",
            "duree_formation": "5 mois",
            "parcours_bta": "Entrepreneur Numérique",
            "secteurs": ["Agences", "Marques", "Médias", "E-commerce", "Politique", "ONG"],
            "perspectives": "Évolution vers Directeur Artistique, UX Designer, Brand Manager.",
            "demande": "Explosion du contenu numérique — chaque entreprise a besoin de contenu quotidien.",
            "freelance": "Excellent — fiverr, 99designs, réseaux sociaux. Revenus très rapides.",
            "competences_futur": ["Design IA", "Motion Design", "AR/VR Design", "3D"],
        },
        "marketeur_digital": {
            "titre": "Marketeur Digital & Growth Hacker",
            "emoji": "📊",
            "description": "Tu développes la présence en ligne des entreprises, gères les publicités et augmentes leurs revenus.",
            "competences": [
                "Meta Ads",
                "Google Ads",
                "SEO",
                "Email Marketing",
                "Analytics",
                "Copywriting",
            ],
            "technologies": [
                "Google Analytics",
                "Facebook Business",
                "HubSpot",
                "Mailchimp",
                "Semrush",
            ],
            "metiers_accessibles": [
                "Traffic Manager",
                "Community Manager",
                "Growth Hacker",
                "CMO Startup",
            ],
            "salaire_haiti": "700-2000",
            "salaire_international": "2500-7000",
            "duree_formation": "8 mois",
            "parcours_bta": "Entrepreneur Numérique",
            "secteurs": ["E-commerce", "Startups", "Agences", "Médias", "Mode", "Tourisme"],
            "perspectives": "Évolution vers CMO, Growth Lead, fondateur d'agence digitale.",
            "demande": "Le marketing digital est indispensable pour toutes les entreprises modernes.",
            "freelance": "Excellent — gestion des réseaux sociaux et publicités pour les PME locales.",
            "competences_futur": ["Marketing IA", "Automatisation", "Créateurs de contenu IA"],
        },
        "cybersecurite": {
            "titre": "Spécialiste en Cybersécurité",
            "emoji": "🔐",
            "description": "Tu protèges les systèmes informatiques contre les cyberattaques et sécurises les données sensibles.",
            "competences": ["Sécurité réseau", "Pentest", "SIEM", "Cryptographie", "Forensique"],
            "technologies": ["Wireshark", "Metasploit", "Nmap", "Kali Linux", "Splunk"],
            "metiers_accessibles": [
                "Analyste SOC",
                "Pentester",
                "RSSI Junior",
                "Consultant Sécurité",
            ],
            "salaire_haiti": "1000-3000",
            "salaire_international": "4000-12000",
            "duree_formation": "8 mois",
            "parcours_bta": "Technicien Informatique Professionnel",
            "secteurs": ["Banques", "Gouvernement", "Défense", "Santé", "Remote"],
            "perspectives": "L'un des métiers les mieux payés du numérique avec pénurie mondiale de talents.",
            "demande": "+35% de croissance annuelle. Les cyberattaques coûtent des milliards aux organisations.",
            "freelance": "Possible — audits de sécurité, tests de pénétration pour les PME.",
            "competences_futur": ["IA en cybersécurité", "Zero Trust", "Cloud Security"],
        },
    }

    profils = [
        ("lyceen_etudiant", "Lycéen / Étudiant", "🎓"),
        ("professionnel", "Professionnel en reconversion", "💼"),
        ("entrepreneur", "Entrepreneur", "🚀"),
        ("sans_emploi", "En recherche d'emploi", "🔍"),
    ]

    interets = [
        ("creer", "Créer des choses (sites, apps, designs)", "🎨"),
        ("analyser", "Analyser et comprendre les données", "📊"),
        ("reparer", "Réparer et configurer des systèmes", "🔧"),
        ("vendre", "Vendre et convaincre", "📢"),
        ("proteger", "Sécuriser et protéger", "🔐"),
        ("automatiser", "Automatiser avec l'IA", "🤖"),
    ]

    objectifs = [
        ("entreprise", "Travailler en entreprise", "🏢"),
        ("freelance", "Travailler en freelance", "💻"),
        ("remote", "Travailler à distance (remote)", "🌍"),
        ("startup", "Créer mon entreprise", "🚀"),
    ]

    niveaux = [
        ("debutant", "Débutant complet", "🌱"),
        ("quelques_bases", "J'ai quelques bases", "📖"),
        ("intermediaire", "Niveau intermédiaire", "⚡"),
    ]

    resultat = None
    metier_data = None
    erreur = None
    form_data = {}

    if request.method == "POST":
        profil = request.POST.get("profil", "")
        interet = request.POST.get("interet", "")
        objectif = request.POST.get("objectif", "")
        niveau = request.POST.get("niveau", "")
        details = request.POST.get("details", "").strip()
        form_data = request.POST

        if profil and interet and objectif and niveau:
            formations = Formation.objects.filter(actif=True).select_related("ecole")

            from .ia import generer_parcours_oriente

            resultat_ia = generer_parcours_oriente(
                profil=profil,
                objectif=f"interet:{interet}, objectif:{objectif}",
                disponibilite=niveau,
                details=details,
                formations_disponibles=formations,
            )

            # Associe le métier correspondant
            mapping_metier = {
                "creer": "developpeur_web",
                "analyser": "expert_ia",
                "reparer": "technicien_informatique",
                "vendre": "marketeur_digital",
                "proteger": "cybersecurite",
                "automatiser": "expert_ia",
            }

            metier_key = mapping_metier.get(interet, "developpeur_web")

            if objectif == "freelance":
                if interet == "creer":
                    metier_key = "designer_graphique"

            metier_data = METIERS_BTA.get(metier_key, METIERS_BTA["developpeur_web"])

            if "erreur" in resultat_ia:
                erreur = resultat_ia.get("erreur")
            else:
                resultat = resultat_ia

        else:
            erreur = "Réponds à toutes les questions pour obtenir ta recommandation."

    return render(
        request,
        "academie/simulateur.html",
        {
            "profils": profils,
            "interets": interets,
            "objectifs": objectifs,
            "niveaux": niveaux,
            "resultat": resultat,
            "metier_data": metier_data,
            "erreur": erreur,
            "form_data": form_data,
        },
    )


# ================================================
# Espace Recrutement / Portfolio
# ================================================


def espace_recrutement(request):
    """Page publique présentant les meilleurs étudiants avec leurs portfolios."""
    from django.contrib.auth.models import User
    from django.db.models import Count, Q

    from .models import BadgeForum, Formation, ProjetEtudiant

    etudiants_qs = (
        User.objects.annotate(
            nb_formations=Count(
                "progressions__lecon__module__formation",
                filter=Q(progressions__terminee=True),
                distinct=True,
            ),
            nb_quiz=Count("resultats_quiz", distinct=True),
            nb_projets=Count("projets", distinct=True),
            nb_badges=Count("badges_forum", distinct=True),
        )
        .filter(Q(nb_formations__gt=0) | Q(nb_projets__gt=0))
        .order_by("-nb_badges", "-nb_formations")[:20]
    )

    etudiants_data = []
    for user in etudiants_qs:
        # Formations complétées à 100%
        formations_completees = []
        for formation in Formation.objects.filter(actif=True):
            if formation.progression_pour(user) == 100:
                formations_completees.append(formation)

        # Projets récents (max 3)
        projets = ProjetEtudiant.objects.filter(auteur=user).order_by("-date_creation")[:3]

        # Badges
        badges = BadgeForum.objects.filter(utilisateur=user)

        etudiants_data.append(
            {
                "user": user,
                "certifications": formations_completees,
                "badges": badges,
                "projets": projets,
                "nb_formations": user.nb_formations,
                "nb_quiz": user.nb_quiz,
                "nb_projets": user.nb_projets,
                "nb_badges": user.nb_badges,
            }
        )

    return render(
        request,
        "academie/recrutement.html",
        {
            "etudiants_data": etudiants_data,
        },
    )


@login_required(login_url="/connexion/")
def mon_portfolio(request):
    """Page où l'étudiant gère ses projets."""
    if request.method == "POST":
        titre = request.POST.get("titre", "").strip()
        description = request.POST.get("description", "").strip()
        technologies = request.POST.get("technologies", "").strip()
        lien = request.POST.get("lien", "").strip()
        image = request.FILES.get("image")

        if titre and description:
            # Vérification du type d'image si présent
            if image:
                kind = filetype.guess(image)
        if kind is None or kind.mime not in ["image/jpeg", "image/png", "image/gif"]:
            messages.error(request, "❌ Format d'image non autorisé. Utilisez JPEG, PNG ou GIF.")
            return redirect("mon_portfolio")

        ProjetEtudiant.objects.create(
            auteur=request.user,
            titre=titre,
            description=description,
            technologies=technologies,
            lien=lien if lien else None,
            image=image,
        )

        messages.success(request, "✅ Projet ajouté avec succès !")
        return redirect("mon_portfolio")
    else:
        messages.error(request, "❌ Titre et description sont obligatoires.")

    projets = ProjetEtudiant.objects.filter(auteur=request.user)
    return render(
        request,
        "academie/portfolio.html",
        {
            "projets": projets,
        },
    )


def verifier_certificat(request, numero):
    """Page publique de vérification d'un certificat."""
    from .models import Certificat

    certificat = None
    try:
        certificat = Certificat.objects.select_related("utilisateur", "formation").get(
            numero=numero
        )
    except Exception:
        pass

    return render(request, "academie/verifier_certificat.html", {"certificat": certificat})


@login_required(login_url="/connexion/")
def notifications_liste(request):
    """Page listant les notifications de l'utilisateur connecté."""
    from .models import Notification

    notifs = Notification.objects.filter(utilisateur=request.user).order_by("-date_creation")[:30]
    # Marque comme lues toutes les notifications affichées
    ids_non_lues = [n.id for n in notifs if not n.lue]
    if ids_non_lues:
        Notification.objects.filter(id__in=ids_non_lues).update(lue=True)
    return render(request, "academie/notifications.html", {"notifications": notifs})


def classement(request):
    """Classement des meilleurs étudiants par XP."""
    from .models import ProfilUtilisateur

    profils = (
        ProfilUtilisateur.objects.select_related("utilisateur")
        .filter(xp__gt=0)
        .order_by("-xp", "-streak")[:50]
    )

    return render(
        request,
        "academie/classement.html",
        {
            "profils": profils,
        },
    )


def set_lang_fr(request):
    """Passe la langue en français et redirige."""

    translation.activate("fr")
    response = redirect(request.META.get("HTTP_REFERER", "/"))
    response.set_cookie(settings.LANGUAGE_COOKIE_NAME, "fr")
    return response


def set_lang_ht(request):
    """Passe la langue en créole et redirige."""

    translation.activate("ht")
    response = redirect(request.META.get("HTTP_REFERER", "/"))
    response.set_cookie(settings.LANGUAGE_COOKIE_NAME, "ht")
    return response


# ================================================
# Page Ressources
# ================================================


def ressources(request):
    """Page principale des ressources — Articles, Outils, Témoignages."""
    categorie = request.GET.get("categorie", "")

    articles = Article.objects.filter(publie=True)
    # === Filtrage par Académie (multi-tenant) ===
    if hasattr(request, "academie_courante") and request.academie_courante:
        articles = articles.filter(Q(academie=request.academie_courante) | Q(academie__isnull=True))
    if categorie:
        articles = articles.filter(categorie=categorie)

    articles_vedette = Article.objects.filter(publie=True, en_vedette=True)[:3]
    outils = OutilRecommande.objects.all()
    temoignages = Temoignage.objects.filter(approuve=True)

    return render(
        request,
        "academie/ressources.html",
        {
            "articles": articles,
            "articles_vedette": articles_vedette,
            "outils": outils,
            "temoignages": temoignages,
            "categorie_active": categorie,
            "categories": Article.CATEGORIES,
            "categories_outils": OutilRecommande.CATEGORIES,
        },
    )


def detail_article(request, slug):
    """Page de détail d'un article."""
    article = Article.objects.get(slug=slug, publie=True)

    articles_lies = Article.objects.filter(publie=True, categorie=article.categorie).exclude(
        id=article.id
    )[:3]

    return render(
        request,
        "academie/detail_article.html",
        {
            "article": article,
            "articles_lies": articles_lies,
        },
    )


# ================================================
# API — Générer article via IA (marketing, resp académique, admin)
# ================================================
@login_required
@role_required("marketing", "resp_academique", "admin", "super_admin")
def api_generer_article(request):
    """API pour générer un article via l'IA."""
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            titre = data.get("titre", "").strip()
            tags = data.get("tags", "").strip()

            if not titre:
                return JsonResponse({"erreur": "Titre requis"}, status=400)

            from .ia import generer_article

            resultat = generer_article(titre, tags)
            return JsonResponse(resultat)
        except Exception as e:
            return JsonResponse({"erreur": str(e)}, status=500)
    return JsonResponse({"erreur": "Méthode non autorisée"}, status=405)


# Pacours Professionnels
def parcours_professionnels(request):
    """Page dédiée aux parcours professionnels — storytelling premium."""
    parcours_list = (
        Parcours.objects.prefetch_related("formations__modules")
        .filter(actif=True)
        .order_by("ordre")
    )

    return render(
        request,
        "academie/parcours.html",
        {
            "parcours_list": parcours_list,
        },
    )


# ================================================
# API — Assistant Back Office (équipe admin)
# ================================================
@extend_schema(
    tags=["Back Office"],
    description="Assistant IA pour les administrateurs – réponse en HTML",
    request=inline_serializer(
        name="AssistantBackOfficeRequest", fields={"question": serializers.CharField()}
    ),
    responses={
        200: inline_serializer(
            name="AssistantBackOfficeResponse", fields={"reponse": serializers.CharField()}
        )
    },
)
@api_view(["POST"])
@permission_classes([IsAuthenticated])
@login_required
@role_required(
    "formateur",
    "resp_academique",
    "marketing",
    "support",
    "finance",
    "direction",
    "admin",
    "super_admin",
)
def api_assistant_backoffice(request):
    """Chat assistant intégré au Back Office."""
    question = request.data.get("question", "").strip()
    if not question:
        return Response({"erreur": "Question vide"}, status=400)

    contexte = (
        f"Rôle: {request.user.profil.get_role_display()}, Utilisateur: {request.user.username}"
    )

    from .ia import assistant_backoffice_ia

    reponse_brute = assistant_backoffice_ia(question, contexte)

    # Convertit la réponse Markdown en HTML pour une meilleure lisibilité
    try:
        reponse_html = markdown_lib.markdown(reponse_brute)
    except Exception:
        reponse_html = reponse_brute  # fallback en cas d'échec

    return Response({"reponse": reponse_html})


@require_POST
def api_simuler_carriere(request):
    """API pour le simulateur de carrière (requêtes AJAX depuis le chatbot)."""
    try:
        data = json.loads(request.body)
        metier = data.get("metier", "").strip()
        if not metier:
            return JsonResponse({"erreur": 'Le champ "metier" est requis.'}, status=400)
        reponse = simuler_carriere_ia(metier=metier)  # ← fonction IA dans ia.py
        return JsonResponse({"reponse": reponse, "metier": metier})
    except json.JSONDecodeError:
        return JsonResponse({"erreur": "JSON invalide"}, status=400)
    except Exception as e:
        return JsonResponse({"erreur": str(e)}, status=500)


# ================================================
# VUE — Aperçu article admin (marketing, resp académique, admin)
# ================================================
@login_required
@role_required("marketing", "resp_academique", "admin", "super_admin")
def apercu_article_admin(request, article_id):
    """Prévisualisation d'un article — brouillon inclus."""
    article = Article.objects.get(id=article_id)
    return render(
        request,
        "academie/detail_article.html",
        {
            "article": article,
            "articles_lies": Article.objects.filter(
                publie=True, categorie=article.categorie
            ).exclude(id=article.id)[:3],
            "mode_apercu": True,
        },
    )


# Vues pour le sitemap et robots.txt
def sitemap_xml(request):
    articles = Article.objects.filter(publie=True)
    formations = Formation.objects.filter(actif=True, slug__isnull=False)
    return render(request, 'academie/sitemap.xml', {
        'articles': articles,
        'formations': formations,
    }, content_type='application/xml')


def robots_txt(request):
    contenu = "User-agent: *\nAllow: /\nSitemap: " + request.build_absolute_uri("/sitemap.xml")
    return HttpResponse(contenu, content_type="text/plain")


# =========================================================
# VUES ADMIN — Prévisualisation et test d'envoi d'emails
# ==========================================================
@login_required
@role_required("resp_academique", "admin", "super_admin")
def admin_email_preview(request, template_name):
    """Prévisualise un email dans le navigateur avec des données factices."""
    contextes_demo = {
        "welcome": {"prenom": "Jean Raymond", "lien_dashboard": "#"},
        "certificate": {
            "prenom": "Jean Raymond",
            "formation_nom": "Python & Django",
            "lien_certificat": "#",
        },
        "badge": {
            "prenom": "Jean Raymond",
            "badge_nom": "Premier Post",
            "badge_icone": "✍️",
            "lien_classement": "#",
        },
        "quiz_result": {
            "prenom": "Jean Raymond",
            "quiz_titre": "Bases de Python",
            "score_texte": "8/10",
            "pourcentage_texte": "80%",
            "message_feedback": "🎉 Excellent travail !",
            "lien_formation": "#",
        },
        "reset_password": {"prenom": "Jean Raymond", "lien_reset": "#"},
        "forum_reply": {
            "auteur_reponse": "Marc B.",
            "sujet_titre": "Comment installer Django ?",
            "extrait_reponse": "Il faut d'abord installer Python...",
            "lien_sujet": "#",
        },
    }
    contexte = contextes_demo.get(template_name, {})
    return render(request, f"emails/notifications/{template_name}.html", contexte)


# ================================================
# VUE — Test email admin (marketing, direction, admin)
# ================================================
@login_required
@role_required("marketing", "direction", "admin", "super_admin")
def admin_email_test(request):
    """Envoie un email de test à l'admin connecté."""
    if request.method == "POST":
        template_name = request.POST.get("template")
        from .email_service import _envoyer_email

        contextes_demo = {
            "welcome": {"prenom": request.user.first_name or "Testeur", "lien_dashboard": "#"},
        }
        _envoyer_email(
            f"emails/notifications/{template_name}.html",
            contextes_demo.get(template_name, {}),
            destinataire=request.user.email,
            sujet=f"[TEST] {template_name}",
        )
        messages.success(
            request, f"✅ Email de test '{template_name}' envoyé à {request.user.email}"
        )
    return redirect("/admin/emails/")


# ================================================
# VUES ADMIN — Centre d'administration des Emails
# ================================================
@login_required
@role_required("marketing", "direction", "admin", "super_admin")
def admin_emails_dashboard(request):
    return render(
        request,
        "admin/emails.html",
        {
            "title": "📧 Emails",
            "site_header": "Blessy Tech Academy",
        },
    )


# =========================================================
# Synchronisation de Contenu
# Interface Admin "Synchronisation"
# Vue Django pour déclencher les commandes
# =========================================================


# ================================================
# VUE — Export synchronisation (direction, admin)
# ================================================
@login_required
@role_required("direction", "admin", "super_admin")
def admin_sync_export(request):
    if request.method == "POST":
        import json
        import time

        from django.http import HttpResponse

        from academie.models import Ecole, Formation, Lecon, Module

        data = {
            "ecoles": [],
            "formations": [],
            "modules": [],
            "lecons": [],
        }

        for ecole in Ecole.objects.all():
            data["ecoles"].append(
                {
                    "nom": ecole.nom,
                    "icone": ecole.icone,
                    "description": ecole.description,
                    "ordre": ecole.ordre,
                }
            )

        for formation in Formation.objects.select_related("ecole").all():
            data["formations"].append(
                {
                    "nom": formation.nom,
                    "ecole": formation.ecole.nom if formation.ecole else None,
                    "icone": formation.icone,
                    "description": formation.description,
                    "duree_mois": formation.duree_mois,
                    "prix": formation.prix,
                    "niveau": formation.niveau,
                    "actif": formation.actif,
                    "gratuit": formation.gratuit,
                }
            )

        for module in Module.objects.select_related("formation").all():
            data["modules"].append(
                {
                    "titre": module.titre,
                    "formation": module.formation.nom,
                    "ordre": module.ordre,
                }
            )

        for lecon in Lecon.objects.select_related("module").all():
            data["lecons"].append(
                {
                    "titre": lecon.titre,
                    "module": lecon.module.titre if lecon.module else None,
                    "contenu": lecon.contenu,
                    "ordre": lecon.ordre,
                }
            )

        nom_fichier = f"bta_export_{time.strftime('%Y%m%d_%H%M%S')}.json"

        response = HttpResponse(
            json.dumps(data, ensure_ascii=False, indent=2), content_type="application/json"
        )
        response["Content-Disposition"] = f'attachment; filename="{nom_fichier}"'
        messages.success(request, f"✅ Export termine : {nom_fichier}")
        return response

    return redirect("/admin/synchronisation/")


# ================================================
# VUE — Import synchronisation (direction, admin)
# ================================================
@login_required
@role_required("direction", "admin", "super_admin")
def admin_sync_import(request):
    if request.method == "POST":
        fichier = request.FILES.get("fichier_import")
        if fichier:
            # Protection contre le Path Traversal
            nom_secure = os.path.basename(fichier.name)
            chemin_temp = os.path.join("/tmp", nom_secure)
            with open(chemin_temp, "wb+") as dest:
                for chunk in fichier.chunks():
                    dest.write(chunk)
            output = StringIO()
            call_command("import_content", chemin_temp, stdout=output)
            messages.success(request, "✅ Import terminé. " + output.getvalue().replace("\n", " "))
            # Nettoyage du fichier temporaire
            os.remove(chemin_temp)
        else:
            messages.error(request, "❌ Aucun fichier fourni.")
    return redirect("/admin/synchronisation/")


# ================================================
# VUE — Dashboard synchronisation (direction, admin)
# ================================================
@login_required
@role_required("direction", "admin", "super_admin")
def admin_sync_dashboard(request):
    from academie.models import Formation

    formations_liste = Formation.objects.filter(actif=True).order_by("nom")
    return render(
        request,
        "admin/synchronisation.html",
        {
            "title": "Synchronisation de contenu",
            "formations_liste": formations_liste,
        },
    )


# ================================================
# Workspace Formation enrichi avec onglets (Dashboard, Stats, Ventes)
# ================================================


@login_required
@role_required("resp_academique", "admin", "super_admin")
def workspace_formation(request, formation_id):
    """Centre de gestion pédagogique complet — tabs + arbre + panneau."""
    formation = get_object_or_404(
        Formation.objects.prefetch_related("modules__lecons", "quiz_set").select_related(
            "ecole", "ecole__academie", "workflow"
        ),
        id=formation_id,
    )

    onglet_actif = request.GET.get("onglet", "dashboard")

    # ---- Données pour l'onglet Dashboard ----
    total_lecons = sum(m.lecons.count() for m in formation.modules.all())
    lecons_avec_contenu = sum(
        1 for m in formation.modules.all() for lecon in m.lecons.all() if lecon.contenu
    )

    # ---- Données pour l'onglet Statistiques ----
    nb_inscrits = 0
    progression_moyenne = 0
    try:
        nb_inscrits = AccesFormationDebloque.objects.filter(formation=formation).count()
        if nb_inscrits > 0:
            etudiants = User.objects.filter(acces_debloques__formation=formation).distinct()
            total_pct = sum(formation.progression_pour(u) for u in etudiants)
            progression_moyenne = round(total_pct / nb_inscrits)
    except NameError:
        pass

    # ---- Données pour l'onglet Vente ----
    ca_formation = 0
    try:
        ca_formation = (
            OrderItem.objects.filter(formation=formation, commande__statut="paye").aggregate(
                t=Sum("prix_unitaire")
            )["t"]
            or 0
        )
    except NameError:
        pass

    # ---- Quiz et examens liés (existants) ----
    quiz_liste = (
        formation.quiz_set.all()
        if hasattr(formation, "quiz_set")
        else Quiz.objects.filter(formation=formation)
    )
    examens = []
    try:
        examens = Examen.objects.filter(formation=formation)
    except NameError:
        pass

    return render(
        request,
        "admin/workspace_formation.html",
        {
            "formation": formation,
            "onglet_actif": onglet_actif,
            "quiz_liste": quiz_liste,
            "examens": examens,
            "total_lecons": total_lecons,
            "lecons_avec_contenu": lecons_avec_contenu,
            "nb_inscrits": nb_inscrits,
            "progression_moyenne": progression_moyenne,
            "ca_formation": ca_formation,
            "title": f"Workspace — {formation.nom}",
            "site_header": admin.site.site_header,
        },
    )


# ================================================
# VUES — Actions Workflow Formation (admin, formateur)
# ================================================


@login_required
@role_required("formateur", "resp_academique", "admin", "super_admin")
def transitionner_workflow_formation(request, formation_id):
    """Action de transition d'état — appelée depuis le Workspace Formation."""
    if request.method == "POST":
        formation = get_object_or_404(Formation, id=formation_id)
        workflow, _ = WorkflowFormation.objects.get_or_create(formation=formation)

        nouvel_etat = request.POST.get("nouvel_etat")
        commentaire = request.POST.get("commentaire", "")

        succes, message = workflow.transitionner(nouvel_etat, request.user, commentaire)

        if succes:
            messages.success(request, f"✅ {message}")
        else:
            messages.error(request, f"❌ {message}")

    return redirect(f"/admin/formation/{formation_id}/workspace/")


@login_required
@role_required("formateur", "resp_academique", "admin", "super_admin")
def mettre_a_jour_checklist(request, formation_id):
    """Coche/décoche un item de checklist qualité."""
    if request.method == "POST":
        formation = get_object_or_404(Formation, id=formation_id)
        workflow, _ = WorkflowFormation.objects.get_or_create(formation=formation)

        champ = request.POST.get("champ")
        valeur = request.POST.get("valeur") == "true"

        if champ in [
            "checklist_contenu_complet",
            "checklist_seo_complet",
            "checklist_prix_valide",
            "checklist_quiz_present",
        ]:
            setattr(workflow, champ, valeur)
            workflow.save()

    return redirect(f"/admin/formation/{formation_id}/workspace/")


# ================================================
# VIEWS.PY — Payment Center
# ================================================


def _prix_avec_promotion(formation):
    """Calcule le prix réel en tenant compte des promotions actives — dynamique, zéro duplication."""
    prix_original = Decimal(str(formation.prix))
    promo_active = Promotion.objects.filter(actif=True).first()
    for promo in Promotion.objects.filter(actif=True):
        if promo.s_applique_a(formation):
            reduction = prix_original * (Decimal(promo.pourcentage_reduction) / 100)
            return prix_original - reduction, promo
    return prix_original, None


@login_required(login_url="/connexion/")
def initier_achat(request, formation_id):
    """Étape 1 — Crée une commande en attente pour une formation."""
    formation = Formation.objects.get(id=formation_id, actif=True)

    if formation.gratuit:
        messages.info(request, "Cette formation est gratuite — accès direct !")
        return redirect("detail_formation", formation_id=formation.id)

    # Empêche le rachat si déjà débloquée
    deja_debloquee = AccesFormationDebloque.objects.filter(
        utilisateur=request.user, formation=formation
    ).exists()
    if deja_debloquee:
        messages.info(request, "Tu as déjà accès à cette formation !")
        return redirect("detail_formation", formation_id=formation.id)

    prix_final, promo = _prix_avec_promotion(formation)

    with db_transaction.atomic():
        commande = Order.objects.create(
            utilisateur=request.user, sous_total=prix_final, total=prix_final
        )
        OrderItem.objects.create(
            commande=commande,
            formation=formation,
            type_produit="formation",
            nom_produit_snapshot=formation.nom,
            icone_produit_snapshot=formation.icone,
            ecole_nom_snapshot=str(formation.ecole) if formation.ecole else "",
            prix_unitaire=prix_final,
        )
        commande.recalculer_total()

    return redirect("checkout", order_reference=commande.reference)


@login_required(login_url="/connexion/")
def checkout(request, order_reference):
    """Étape 2 — Page de paiement : choix moyen + coupon."""
    commande = Order.objects.prefetch_related("items").get(
        reference=order_reference, utilisateur=request.user
    )

    if request.method == "POST":
        code_coupon = request.POST.get("code_coupon", "").strip().upper()
        moyen_id = request.POST.get("moyen_paiement")

        if code_coupon:
            try:
                coupon = Coupon.objects.get(code=code_coupon)
                valide, message_erreur = coupon.est_valide()
                if valide:
                    commande.coupon_applique = coupon
                    commande.save()
                    commande.recalculer_total()
                    messages.success(request, f"✅ Coupon '{code_coupon}' appliqué !")
                else:
                    messages.error(request, f"❌ {message_erreur}")
            except Coupon.DoesNotExist:
                messages.error(request, "❌ Code coupon invalide.")
            return redirect("checkout", order_reference=order_reference)

        if moyen_id:
            moyen = MoyenPaiement.objects.get(id=moyen_id)
            commande.moyen_paiement = moyen
            commande.save()
            return redirect("confirmer_paiement", order_reference=order_reference)

    moyens_paiement = MoyenPaiement.objects.filter(actif=True)
    return render(
        request,
        "academie/checkout.html",
        {
            "commande": commande,
            "moyens_paiement": moyens_paiement,
        },
    )


@login_required(login_url="/connexion/")
def confirmer_paiement(request, order_reference):
    """
    Étape 3 — Confirmation.
    Paiement manuel = upload preuve + statut "en_verification" (admin valide ensuite).
    """
    commande = Order.objects.get(reference=order_reference, utilisateur=request.user)

    if request.method == "POST":
        preuve = request.FILES.get("preuve_paiement")
        reference_externe = request.POST.get("reference_externe", "")

        with db_transaction.atomic():
            Transaction.objects.create(
                commande=commande,
                moyen_paiement=commande.moyen_paiement,
                reference_externe=reference_externe,
                preuve_paiement=preuve,
                montant=commande.total,
                statut="en_verification" if commande.moyen_paiement.code == "manuel" else "initiee",
            )

        messages.success(
            request,
            "✅ Paiement soumis ! Notre équipe valide généralement sous 24h. "
            "Tu recevras un email de confirmation dès validation.",
        )
        return redirect("mes_commandes")

    return render(request, "academie/confirmer_paiement.html", {"commande": commande})


# ================================================
# VUE — Valider transaction (finance, admin)
# ================================================
@login_required
@role_required("finance", "admin", "super_admin")
def admin_valider_transaction(request, transaction_id):
    """Admin — valide manuellement un paiement et débloque l'accès automatiquement."""
    trans = Transaction.objects.select_related("commande").get(id=transaction_id)
    with db_transaction.atomic():
        trans.statut = "reussie"
        trans.valide_par = request.user
        trans.save()

        commande = trans.commande
        commande.statut = "paye"
        commande.date_paiement = timezone.now()
        commande.save()

        # Débloque l'accès pour CHAQUE item de la commande
        for item in commande.items.all():
            if item.formation:
                AccesFormationDebloque.objects.get_or_create(
                    utilisateur=commande.utilisateur,
                    nom_formation_snapshot=item.nom_produit_snapshot,
                    defaults={"formation": item.formation, "commande_origine": commande},
                )

        # Génère la facture automatiquement
        facture, _ = Invoice.objects.get_or_create(commande=commande)

        # Incrémente l'usage du coupon si utilisé
        if commande.coupon_applique:
            commande.coupon_applique.utilisations_actuelles += 1
            commande.coupon_applique.save()

        # Envoie email de confirmation (réutilise email_service.py déjà construit)
        try:
            from .email_service import _envoyer_email

            _envoyer_email(
                "emails/notifications/payment_confirmed.html",
                {
                    "prenom": commande.utilisateur.first_name or commande.utilisateur.username,
                    "commande": commande,
                    "facture_numero": facture.numero_facture,
                },
                destinataire=commande.utilisateur.email,
                sujet=f"✅ Paiement confirmé — Commande {commande.reference}",
            )
        except Exception:
            pass

    # Journalisation (hors transaction pour ne pas bloquer)
    enregistrer_log(
        request,
        "validation_paiement",
        f"Transaction {trans.id} validée pour commande {commande.reference} ({commande.total}$)",
        "Transaction",
        trans.id,
    )

    messages.success(
        request, f"✅ Transaction validée — Accès débloqué pour {commande.utilisateur.username}"
    )
    return redirect("/admin/academie/transaction/")


@login_required(login_url="/connexion/")
def mes_commandes(request):
    """Dashboard étudiant — Mes commandes/factures/remboursements."""
    commandes = Order.objects.filter(utilisateur=request.user).prefetch_related("items", "facture")
    return render(request, "academie/mes_commandes.html", {"commandes": commandes})


def verifier_acces_formation(user, formation):
    """
    Fonction utilitaire — à réutiliser PARTOUT où on doit vérifier
    si un étudiant a accès à une formation payante.
    Usage : if verifier_acces_formation(request.user, formation): ...
    """
    if formation.gratuit:
        return True
    if not user.is_authenticated:
        return False
    return AccesFormationDebloque.objects.filter(utilisateur=user, formation=formation).exists()


# ================================================
# VIEWS.PY — Intégration des passerelles réelles
# Remplace/complète confirmer_paiement() selon moyen choisi
# ================================================


@login_required(login_url="/connexion/")
def rediriger_paiement_externe(request, order_reference):
    """Route vers la bonne passerelle selon le moyen choisi."""
    commande = Order.objects.get(reference=order_reference, utilisateur=request.user)
    code_moyen = commande.moyen_paiement.code

    url_succes = request.build_absolute_uri(f"/paiement-succes/{order_reference}/")
    url_annulation = request.build_absolute_uri(f"/checkout/{order_reference}/")

    if code_moyen == "stripe":
        from academie.payment_gateways import stripe_gateway

        url, session_id = stripe_gateway.creer_session_paiement(
            commande, url_succes, url_annulation
        )
        if url:
            Transaction.objects.create(
                commande=commande,
                moyen_paiement=commande.moyen_paiement,
                reference_externe=session_id,
                montant=commande.total,
                statut="initiee",
            )
            return redirect(url)
        messages.error(request, f"Erreur Stripe : {session_id}")

    elif code_moyen == "moncash":
        from academie.payment_gateways import moncash_gateway

        url, erreur = moncash_gateway.creer_paiement(commande)
        if url:
            return redirect(url)
        messages.error(request, erreur)

    elif code_moyen == "paypal":
        from academie.payment_gateways import paypal_gateway

        url, payment_id = paypal_gateway.creer_paiement(commande, url_succes, url_annulation)
        if url:
            Transaction.objects.create(
                commande=commande,
                moyen_paiement=commande.moyen_paiement,
                reference_externe=payment_id,
                montant=commande.total,
                statut="initiee",
            )
            return redirect(url)
        messages.error(request, "Erreur PayPal")

    return redirect("confirmer_paiement", order_reference=order_reference)


@login_required(login_url="/connexion/")
def paiement_succes(request, order_reference):
    """Page de retour après paiement externe réussi — débloque l'accès."""
    commande = Order.objects.get(reference=order_reference, utilisateur=request.user)

    with db_transaction.atomic():
        commande.statut = "paye"
        commande.date_paiement = timezone.now()
        commande.save()

        for item in commande.items.all():
            if item.formation:
                AccesFormationDebloque.objects.get_or_create(
                    utilisateur=commande.utilisateur,
                    nom_formation_snapshot=item.nom_produit_snapshot,
                    defaults={"formation": item.formation, "commande_origine": commande},
                )
        Invoice.objects.get_or_create(commande=commande)

    messages.success(request, "🎉 Paiement confirmé ! Accès débloqué immédiatement.")
    return redirect("mes_commandes")


@csrf_exempt
def stripe_webhook(request):
    """Endpoint webhook Stripe — confirmation asynchrone officielle."""
    from academie.payment_gateways import stripe_gateway

    event = stripe_gateway.traiter_webhook(request.body, request.META.get("HTTP_STRIPE_SIGNATURE"))

    if event and event["type"] == "checkout.session.completed":
        session = event["data"]["object"]
        reference = session.get("client_reference_id")
        try:
            commande = Order.objects.get(reference=reference)
            if commande.statut != "paye":
                commande.statut = "paye"
                commande.date_paiement = timezone.now()
                commande.save()
                for item in commande.items.all():
                    if item.formation:
                        AccesFormationDebloque.objects.get_or_create(
                            utilisateur=commande.utilisateur,
                            nom_formation_snapshot=item.nom_produit_snapshot,
                            defaults={"formation": item.formation, "commande_origine": commande},
                        )
                Invoice.objects.get_or_create(commande=commande)
        except Order.DoesNotExist:
            pass

    return HttpResponse(status=200)


# ================================================
# VUE — Dashboard Business (finance, direction, admin)
# ================================================
@login_required
@role_required("finance", "direction", "admin", "super_admin")
def vue_dashboard_business(request):
    import json
    from datetime import timedelta

    total_inscriptions = Inscription.objects.count()
    inscriptions_non_traitees = Inscription.objects.filter(traite=False).count()

    # Revenus réels (basés sur commandes payées — pas potentiels)
    ca_total = Order.objects.filter(statut="paye").aggregate(total=Sum("total"))["total"] or 0

    # Ventes des 30 derniers jours (pour graphique)
    labels_jours, valeurs_jours = [], []
    for i in range(29, -1, -1):
        jour = timezone.now().date() - timedelta(days=i)
        montant_jour = (
            Order.objects.filter(statut="paye", date_paiement__date=jour).aggregate(
                total=Sum("total")
            )["total"]
            or 0
        )
        labels_jours.append(jour.strftime("%d/%m"))
        valeurs_jours.append(float(montant_jour))

    # Top formations vendues
    formations_populaires = Formation.objects.annotate(
        nb_ventes=Count("orderitem", filter=Q(orderitem__commande__statut="paye"))
    ).order_by("-nb_ventes")[:8]

    # Répartition par moyen de paiement
    repartition_moyens = (
        Transaction.objects.filter(statut="reussie")
        .values("moyen_paiement__nom_affiche")
        .annotate(total=Count("id"))
    )

    coupons_utilises = Coupon.objects.filter(utilisations_actuelles__gt=0).count()
    remboursements_total = (
        Refund.objects.filter(statut="approuve").aggregate(total=Sum("montant"))["total"] or 0
    )

    return render(
        request,
        "admin/dashboard_business.html",
        {
            "title": "💼 Dashboard Business",
            "site_header": admin.site.site_header,
            "ca_total": ca_total,
            "total_inscriptions": total_inscriptions,
            "inscriptions_non_traitees": inscriptions_non_traitees,
            "formations_populaires": formations_populaires,
            "coupons_utilises": coupons_utilises,
            "remboursements_total": remboursements_total,
            "chart_labels_json": json.dumps(labels_jours),
            "chart_valeurs_json": json.dumps(valeurs_jours),
            "chart_moyens_labels": json.dumps(
                [m["moyen_paiement__nom_affiche"] or "N/A" for m in repartition_moyens]
            ),
            "chart_moyens_valeurs": json.dumps([m["total"] for m in repartition_moyens]),
        },
    )


# ================================================
# VIEWS.PY — Export Excel/PDF des ventes (finance, admin)
# ================================================


@login_required
@role_required("finance", "admin", "super_admin")
def export_ventes_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Ventes BTA"

    entetes = ["Référence", "Client", "Formation", "Montant", "Statut", "Date"]
    ws.append(entetes)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0B2447", fill_type="solid")

    commandes = Order.objects.filter(statut="paye").prefetch_related("items")
    for cmd in commandes:
        for item in cmd.items.all():
            ws.append(
                [
                    cmd.reference,
                    cmd.utilisateur.username,
                    item.nom_produit_snapshot,
                    float(item.prix_unitaire),
                    cmd.get_statut_display(),
                    cmd.date_paiement.strftime("%d/%m/%Y") if cmd.date_paiement else "",
                ]
            )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="ventes_bta.xlsx"'
    wb.save(response)
    return response


@login_required
@role_required("finance", "admin", "super_admin")
def export_ventes_pdf(request):
    """Export des ventes au format PDF."""
    try:
        from weasyprint import HTML
    except (ImportError, OSError):
        messages.warning(request, "📄 L'export PDF n'est pas disponible. Utilisez l'export Excel.")
        return redirect("/admin/dashboard-business/")

    commandes = Order.objects.filter(statut="paye").prefetch_related("items")
    ca_total = commandes.aggregate(total=Sum("total"))["total"] or 0

    html_string = render_to_string(
        "academie/pdf/rapport_ventes.html",
        {
            "commandes": commandes,
            "ca_total": ca_total,
            "date_generation": timezone.now(),
        },
    )
    pdf = HTML(string=html_string).write_pdf()
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="rapport_ventes_bta.pdf"'
    return response


# ================================================
# VUES — Plateforme d'examens officiels modernisée
# ================================================


@login_required
def preparation_examen(request, examen_id):
    """Page de préparation avant examen — vérifications et checklist."""
    examen = get_object_or_404(Examen, id=examen_id, actif=True)

    # Vérifications d'accès
    erreurs = []

    # Vérifier date de disponibilité
    if examen.date_disponibilite and timezone.now() < examen.date_disponibilite:
        erreurs.append(
            f"L'examen sera disponible le {examen.date_disponibilite.strftime('%d/%m/%Y à %H:%M')}"
        )

    # Vérifier date d'expiration
    if examen.date_expiration and timezone.now() > examen.date_expiration:
        erreurs.append("Cet examen n'est plus disponible.")

    # Vérifier tentatives
    tentatives_count = TentativeExamen.objects.filter(
        utilisateur=request.user, examen=examen
    ).count()
    tentatives_restantes = examen.tentatives_max - tentatives_count
    if tentatives_restantes <= 0:
        erreurs.append("Vous avez atteint le nombre maximum de tentatives.")

    # Compétences formatées
    competences = [c.strip() for c in examen.competences_evaluees.split("\n") if c.strip()]
    prerequis = [p.strip() for p in examen.prerequis_examen.split("\n") if p.strip()]
    conditions = (
        [c.strip() for c in examen.conditions_utilisation.split("\n") if c.strip()]
        if examen.conditions_utilisation
        else []
    )

    checklist = [
        ("Connexion Internet stable", "wifi"),
        ("Batterie suffisante ou secteur branché", "battery"),
        ("Navigateur compatible (Chrome, Firefox, Edge)", "browser"),
        ("Être dans un endroit calme, sans distraction", "quiet"),
    ]

    return render(
        request,
        "academie/preparation_examen.html",
        {
            "examen": examen,
            "erreurs": erreurs,
            "tentatives_restantes": tentatives_restantes,
            "competences": competences,
            "prerequis": prerequis,
            "conditions": conditions,
            "checklist": checklist,
        },
    )


@login_required
def passer_examen(request, examen_id):
    """Interface d'examen chronométrée."""
    examen = get_object_or_404(Examen, id=examen_id, actif=True)

    questions = list(examen.questions.prefetch_related("choix").all())
    random.shuffle(questions)

    for question in questions:
        choix_list = list(question.choix.all())
        random.shuffle(choix_list)
        question.choix_melanges = choix_list

    return render(
        request,
        "academie/examen.html",
        {
            "examen": examen,
            "questions": questions,
            "duree_secondes": examen.duree_minutes * 60,
        },
    )


@login_required
def soumettre_examen(request, examen_id):
    """Soumission et correction de l'examen."""
    if request.method != "POST":
        return redirect("passer_examen", examen_id=examen_id)

    examen = get_object_or_404(Examen, id=examen_id, actif=True)

    tentative = TentativeExamen.objects.create(
        utilisateur=request.user,
        examen=examen,
    )

    evenements = request.POST.get("evenements_suspects", "[]")
    try:
        tentative.evenements_suspects = json.loads(evenements)
    except Exception:
        pass

    temps_utilise = request.POST.get("temps_utilise", 0)
    tentative.temps_utilise_secondes = int(temps_utilise) if temps_utilise else 0

    total_points = 0
    points_obtenus = 0
    bonnes = 0
    mauvaises = 0
    repondues = 0

    for question in examen.questions.all():
        total_points += question.points
        reponse_key = f"question_{question.id}"

        if question.type_question == "qcm":
            choix_id = request.POST.get(reponse_key)
            if choix_id:
                repondues += 1
                choix = get_object_or_404(ChoixExamen, id=choix_id)
                if choix.est_correct:
                    points_obtenus += question.points
                    bonnes += 1
                else:
                    mauvaises += 1
        elif question.type_question in ["vrai_faux", "texte"]:
            valeur = request.POST.get(reponse_key)
            if valeur:
                repondues += 1

    score = round((points_obtenus / total_points) * 100, 1) if total_points > 0 else 0
    tentative.score = score
    tentative.reussi = score >= examen.seuil_reussite
    tentative.questions_repondues = repondues
    tentative.bonnes_reponses = bonnes
    tentative.mauvaises_reponses = mauvaises
    tentative.date_fin = timezone.now()
    tentative.save()

    # === Enrichissement post-examen (XP, Certificat, Feedback IA) ===
    if tentative.reussi:
        from .xp_utils import ajouter_xp

        ajouter_xp(request.user, examen.xp_recompense or 50)

        if examen.certificat_auto:
            from .models import Certificat

            Certificat.objects.get_or_create(
                utilisateur=request.user,
                formation=examen.formation,
            )

    if score < 70:
        contexte_feedback = f"L'étudiant a obtenu {score}% à l'examen {examen.titre}. Donne un conseil bref et motivant en 2 phrases."
    else:
        contexte_feedback = f"L'étudiant a réussi l'examen {examen.titre} avec {score}%. Félicite-le brièvement en 2 phrases."

    try:
        from .ia import initialiser_ia

        client = initialiser_ia()
        response = client.models.generate_content(
            model="gemini-2.5-flash", contents=contexte_feedback
        )
        feedback_ia = response.text
    except Exception:
        feedback_ia = (
            "Continue à pratiquer régulièrement — chaque tentative te rapproche de la maîtrise !"
        )

    return render(
        request,
        "academie/resultat_examen.html",
        {
            "examen": examen,
            "tentative": tentative,
            "score": score,
            "reussi": tentative.reussi,
            "feedback_ia": feedback_ia,
        },
    )


# ================================================
# VUES — Dashboard CRM (marketing, support, admin)
# ================================================
@login_required
@role_required("marketing", "support", "admin", "super_admin")
def dashboard_crm(request):
    """Centre CRM — gestion des leads/prospects."""
    statut_filtre = request.GET.get("statut", "")

    leads = Inscription.objects.select_related("formation", "assigne_a").prefetch_related(
        "interactions"
    )
    if statut_filtre:
        leads = leads.filter(statut_lead=statut_filtre)

    stats = {
        "total": Inscription.objects.count(),
        "nouveaux": Inscription.objects.filter(statut_lead="nouveau").count(),
        "convertis": Inscription.objects.filter(statut_lead="converti").count(),
        "perdus": Inscription.objects.filter(statut_lead="perdu").count(),
    }

    return render(
        request,
        "admin/dashboard_crm.html",
        {
            "title": "📢 Centre CRM",
            "leads": leads.order_by("-date_inscription")[:50],
            "stats": stats,
            "statut_filtre": statut_filtre,
            "statuts": Inscription._meta.get_field("statut_lead").choices,
        },
    )


@login_required
@role_required("marketing", "support", "admin", "super_admin")
def ajouter_interaction_crm(request, inscription_id):
    """Ajoute une interaction à un lead."""
    if request.method == "POST":
        inscription = Inscription.objects.get(id=inscription_id)
        InteractionCRM.objects.create(
            inscription=inscription,
            type_interaction=request.POST.get("type_interaction", "note"),
            contenu=request.POST.get("contenu", ""),
            auteur=request.user,
        )
        nouveau_statut = request.POST.get("nouveau_statut")
        if nouveau_statut:
            inscription.statut_lead = nouveau_statut
            inscription.save()
        messages.success(request, "✅ Interaction ajoutée")
    return redirect("dashboard_crm")


# ================================================
# VUE — Dashboard Enseignant (formateur, admin)
# ================================================
@login_required
@role_required("formateur", "resp_academique", "admin", "super_admin")
def dashboard_enseignant(request):
    """Espace personnel du formateur connecté — ses formations et revenus."""
    try:
        enseignant = request.user.profil.enseignant
    except (AttributeError, Enseignant.DoesNotExist):
        messages.error(request, "❌ Aucun profil enseignant associé à ce compte.")
        return redirect("dashboard")

    formations = enseignant.formations_attribuees.all()

    return render(
        request,
        "academie/dashboard_enseignant.html",
        {
            "enseignant": enseignant,
            "formations": formations,
            "revenus": enseignant.revenus_generes(),
            "remuneration": enseignant.part_remuneration(),
            "nb_etudiants": enseignant.nb_etudiants_formes(),
            "note_moyenne": enseignant.note_moyenne_temoignages(),
        },
    )


# ================================================
# VIEWS.PY — Page offline (fallback PWA)
# ================================================


def page_offline(request):
    return render(request, "academie/offline.html")
