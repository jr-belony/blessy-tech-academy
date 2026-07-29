# ================================================
# VIEWS_MODULES/ADMIN_VIEWS.PY — Vues Back Office custom
# ================================================

import json
import os
import markdown as markdown_lib
import logging
from datetime import timedelta
from decimal import Decimal
from io import StringIO

from django.contrib import admin, messages
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core.management import call_command
from django.db import transaction as db_transaction
from django.db.models import Avg, Count, Q, Sum
from django.db.models.functions import TruncMonth
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.utils import timezone
from django.views.decorators.http import require_POST

from drf_spectacular.utils import extend_schema, inline_serializer
from rest_framework import serializers
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from users.models import Enseignant

from ..models import (
    AccesFormationDebloque,
    Article,
    Certificat,
    Coupon,
    DisponibiliteMentor,
    Ecole,
    Examen,
    Formation,
    GradebookEntry,
    Inscription,
    Invoice,
    Lecon,
    LogAudit,
    Module,
    Order,
    OrderItem,
    Parcours,
    Promotion,
    Quiz,
    Refund,
    Reponse,
    ReservationMentorat,
    Sujet,
    Transaction,
    WorkflowFormation,
)
from ..permissions import enregistrer_log, role_required
from ..services.async_tasks import executer_en_arriere_plan
from ..services.email_service import _envoyer_email
from ..services.ia_service import (
    analyser_plateforme_ia,
    assistant_backoffice_ia,
)
from ..xp_utils import ajouter_xp
from .. import notifications
logger = logging.getLogger('academie')


# ================================================
# Workspace Formation
# ================================================

@login_required
@role_required("resp_academique", "admin", "super_admin")
def workspace_formation(request, formation_id):
    """Centre de gestion pédagogique complet — tabs + arbre + panneau."""
    formation = get_object_or_404(
        Formation.objects.prefetch_related("modules__lecons", "quiz_set").select_related(
            "ecole", "ecole__academie", "workflow"
        ),
        id=formation_id,
    )

    onglet_actif = request.GET.get("onglet", "dashboard")

    total_lecons = sum(m.lecons.count() for m in formation.modules.all())
    lecons_avec_contenu = sum(
        1 for m in formation.modules.all() for lecon in m.lecons.all() if lecon.contenu
    )

    nb_inscrits = 0
    progression_moyenne = 0
    try:
        nb_inscrits = AccesFormationDebloque.objects.filter(formation=formation).count()
        if nb_inscrits > 0:
            etudiants = User.objects.filter(acces_debloques__formation=formation).distinct()
            total_pct = sum(formation.progression_pour(u) for u in etudiants)
            progression_moyenne = round(total_pct / nb_inscrits)
    except NameError:
        pass

    ca_formation = 0
    try:
        ca_formation = (
            OrderItem.objects.filter(formation=formation, commande__statut="paye").aggregate(
                t=Sum("prix_unitaire")
            )["t"]
            or 0
        )
    except NameError:
        pass

    quiz_liste = (
        formation.quiz_set.all()
        if hasattr(formation, "quiz_set")
        else Quiz.objects.filter(formation=formation)
    )
    examens = []
    try:
        examens = Examen.objects.filter(formation=formation)
    except NameError:
        pass

    return render(
        request,
        "admin/workspace_formation.html",
        {
            "formation": formation,
            "onglet_actif": onglet_actif,
            "quiz_liste": quiz_liste,
            "examens": examens,
            "total_lecons": total_lecons,
            "lecons_avec_contenu": lecons_avec_contenu,
            "nb_inscrits": nb_inscrits,
            "progression_moyenne": progression_moyenne,
            "ca_formation": ca_formation,
            "title": f"Workspace — {formation.nom}",
            "site_header": admin.site.site_header,
        },
    )


@login_required
@role_required("formateur", "resp_academique", "admin", "super_admin")
def transitionner_workflow_formation(request, formation_id):
    """Action de transition d'état — appelée depuis le Workspace Formation."""
    if request.method == "POST":
        formation = get_object_or_404(Formation, id=formation_id)
        workflow, _ = WorkflowFormation.objects.get_or_create(formation=formation)

        nouvel_etat = request.POST.get("nouvel_etat")
        commentaire = request.POST.get("commentaire", "")

        succes, message = workflow.transitionner(nouvel_etat, request.user, commentaire)

        if succes:
            messages.success(request, f"✅ {message}")
        else:
            messages.error(request, f"❌ {message}")

    return redirect(f"/admin/formation/{formation_id}/workspace/")


@login_required
@role_required("formateur", "resp_academique", "admin", "super_admin")
def mettre_a_jour_checklist(request, formation_id):
    """Coche/décoche un item de checklist qualité."""
    if request.method == "POST":
        formation = get_object_or_404(Formation, id=formation_id)
        workflow, _ = WorkflowFormation.objects.get_or_create(formation=formation)

        champ = request.POST.get("champ")
        valeur = request.POST.get("valeur") == "true"

        if champ in [
            "checklist_contenu_complet",
            "checklist_seo_complet",
            "checklist_prix_valide",
            "checklist_quiz_present",
        ]:
            setattr(workflow, champ, valeur)
            workflow.save()

    return redirect(f"/admin/formation/{formation_id}/workspace/")


# ================================================
# Admin — Validation de transaction
# ================================================

@login_required
@role_required("finance", "admin", "super_admin")
def admin_valider_transaction(request, transaction_id):
    """Admin — valide un paiement et débloque automatiquement les accès."""
    trans = Transaction.objects.select_related("commande").get(id=transaction_id)
    with db_transaction.atomic():
        trans.statut = "reussie"
        trans.valide_par = request.user
        trans.save()
        commande = trans.commande
        commande.statut = "paye"
        commande.date_paiement = timezone.now()
        commande.save()

        for item in commande.items.all():
            if item.formation:
                AccesFormationDebloque.objects.get_or_create(
                    utilisateur=commande.utilisateur,
                    nom_formation_snapshot=item.nom_produit_snapshot,
                    defaults={
                        "formation": item.formation,
                        "commande_origine": commande,
                    },
                )

        facture, _ = Invoice.objects.get_or_create(commande=commande)
        if commande.coupon_applique:
            succes, msg = commande.coupon_applique.utiliser_atomiquement()
            if not succes:
                logger.warning(f"Échec coupon : {msg}")

        try:
            from ..tasks import (
                tache_envoyer_email,
                tache_generer_facture_pdf,
            )
            tache_generer_facture_pdf.send(commande.id)
            tache_envoyer_email.send(
                "emails/notifications/payment_confirmed.html",
                {
                    "prenom": (
                        commande.utilisateur.first_name
                        or commande.utilisateur.username
                    ),
                    "commande": commande,
                    "facture_numero": facture.numero_facture,
                },
                destinataire=commande.utilisateur.email,
                sujet=f"✅ Paiement confirmé — Commande {commande.reference}",
            )
        except Exception:
            executer_en_arriere_plan(
                _envoyer_email,
                "emails/notifications/payment_confirmed.html",
                {
                    "prenom": (
                        commande.utilisateur.first_name
                        or commande.utilisateur.username
                    ),
                    "commande": commande,
                    "facture_numero": facture.numero_facture,
                },
                destinataire=commande.utilisateur.email,
                sujet=f"✅ Paiement confirmé — Commande {commande.reference}",
            )

    enregistrer_log(
        request,
        "validation_paiement",
        f"Transaction {trans.id} validée pour commande {commande.reference} ({commande.total}$)",
        "Transaction",
        trans.id,
    )
    messages.success(
        request,
        f"✅ Transaction validée — Accès débloqué pour {commande.utilisateur.username}",
    )
    return redirect("/admin/academie/transaction/")


# ================================================
# Vues Admin — Emails
# ================================================

@login_required
@role_required("resp_academique", "admin", "super_admin")
def admin_email_preview(request, template_name):
    """Prévisualise un email dans le navigateur avec des données factices."""
    contextes_demo = {
        "welcome": {"prenom": "Jean Raymond", "lien_dashboard": "#"},
        "certificate": {
            "prenom": "Jean Raymond",
            "formation_nom": "Python & Django",
            "lien_certificat": "#",
        },
        "badge": {
            "prenom": "Jean Raymond",
            "badge_nom": "Premier Post",
            "badge_icone": "✍️",
            "lien_classement": "#",
        },
        "quiz_result": {
            "prenom": "Jean Raymond",
            "quiz_titre": "Bases de Python",
            "score_texte": "8/10",
            "pourcentage_texte": "80%",
            "message_feedback": "🎉 Excellent travail !",
            "lien_formation": "#",
        },
        "reset_password": {"prenom": "Jean Raymond", "lien_reset": "#"},
        "forum_reply": {
            "auteur_reponse": "Marc B.",
            "sujet_titre": "Comment installer Django ?",
            "extrait_reponse": "Il faut d'abord installer Python...",
            "lien_sujet": "#",
        },
    }
    contexte = contextes_demo.get(template_name, {})
    return render(request, f"emails/notifications/{template_name}.html", contexte)


@login_required
@role_required("marketing", "direction", "admin", "super_admin")
def admin_email_test(request):
    """Envoie un email de test à l'admin connecté."""
    if request.method == "POST":
        template_name = request.POST.get("template")
        contextes_demo = {
            "welcome": {"prenom": request.user.first_name or "Testeur", "lien_dashboard": "#"},
        }
        _envoyer_email(
            f"emails/notifications/{template_name}.html",
            contextes_demo.get(template_name, {}),
            destinataire=request.user.email,
            sujet=f"[TEST] {template_name}",
        )
        messages.success(
            request, f"✅ Email de test '{template_name}' envoyé à {request.user.email}"
        )
    return redirect("/admin/emails/")


@login_required
@role_required("marketing", "direction", "admin", "super_admin")
def admin_emails_dashboard(request):
    """Centre d'administration des Emails."""
    return render(
        request,
        "admin/emails.html",
        {
            "title": "📧 Emails",
            "site_header": "Blessy Tech Academy",
        },
    )


# ================================================
# Synchronisation
# ================================================

@login_required
@role_required("direction", "admin", "super_admin")
def admin_sync_export(request):
    """Export synchronisation — génère un fichier JSON du contenu."""
    if request.method == "POST":
        import json
        import time

        data = {
            "ecoles": [],
            "formations": [],
            "modules": [],
            "lecons": [],
        }

        for ecole in Ecole.objects.all():
            data["ecoles"].append(
                {
                    "nom": ecole.nom,
                    "icone": ecole.icone,
                    "description": ecole.description,
                    "ordre": ecole.ordre,
                }
            )

        for formation in Formation.objects.select_related("ecole").all():
            data["formations"].append(
                {
                    "nom": formation.nom,
                    "ecole": formation.ecole.nom if formation.ecole else None,
                    "icone": formation.icone,
                    "description": formation.description,
                    "duree_mois": formation.duree_mois,
                    "prix": formation.prix,
                    "niveau": formation.niveau,
                    "actif": formation.actif,
                    "gratuit": formation.gratuit,
                }
            )

        for module in Module.objects.select_related("formation").all():
            data["modules"].append(
                {
                    "titre": module.titre,
                    "formation": module.formation.nom,
                    "ordre": module.ordre,
                }
            )

        for lecon in Lecon.objects.select_related("module").all():
            data["lecons"].append(
                {
                    "titre": lecon.titre,
                    "module": lecon.module.titre if lecon.module else None,
                    "contenu": lecon.contenu,
                    "ordre": lecon.ordre,
                }
            )

        nom_fichier = f"bta_export_{time.strftime('%Y%m%d_%H%M%S')}.json"

        response = HttpResponse(
            json.dumps(data, ensure_ascii=False, indent=2), content_type="application/json"
        )
        response["Content-Disposition"] = f'attachment; filename="{nom_fichier}"'
        messages.success(request, f"✅ Export termine : {nom_fichier}")
        return response

    return redirect("/admin/synchronisation/")


@login_required
@role_required("direction", "admin", "super_admin")
def admin_sync_import(request):
    """Import synchronisation — charge un fichier JSON et exécute la commande import_content."""
    if request.method == "POST":
        fichier = request.FILES.get("fichier_import")
        if fichier:
            nom_secure = os.path.basename(fichier.name)
            chemin_temp = os.path.join("/tmp", nom_secure)
            with open(chemin_temp, "wb+") as dest:
                for chunk in fichier.chunks():
                    dest.write(chunk)
            output = StringIO()
            call_command("import_content", chemin_temp, stdout=output)
            messages.success(request, "✅ Import terminé. " + output.getvalue().replace("\n", " "))
            os.remove(chemin_temp)
        else:
            messages.error(request, "❌ Aucun fichier fourni.")
    return redirect("/admin/synchronisation/")


@login_required
@role_required("direction", "admin", "super_admin")
def admin_sync_dashboard(request):
    """Dashboard synchronisation."""
    formations_liste = Formation.objects.filter(actif=True).order_by("nom")
    return render(
        request,
        "admin/synchronisation.html",
        {
            "title": "Synchronisation de contenu",
            "formations_liste": formations_liste,
        },
    )


# ================================================
# Backup complet depuis l'admin
# ================================================

@staff_member_required
def admin_backup_complet(request):
    """Déclenchement backup complet depuis l'admin."""
    if request.method == "POST":
        output = StringIO()
        try:
            call_command("backup_database", stdout=output)
            messages.success(
                request,
                "✅ Sauvegarde lancée. " + output.getvalue().replace("\n", " ")
            )
        except Exception as e:
            messages.error(
                request,
                f"❌ Erreur pendant la sauvegarde : {str(e)}"
            )
    return redirect("/admin/synchronisation/")


# ================================================
# Aperçu article admin
# ================================================

@login_required
@role_required("marketing", "resp_academique", "admin", "super_admin")
def apercu_article_admin(request, article_id):
    """Prévisualisation d'un article — brouillon inclus."""
    article = Article.objects.get(id=article_id)
    return render(
        request,
        "academie/detail_article.html",
        {
            "article": article,
            "articles_lies": Article.objects.filter(
                publie=True, categorie=article.categorie
            ).exclude(id=article.id)[:3],
            "mode_apercu": True,
        },
    )


# ================================================
# Export Excel/PDF des ventes
# ================================================

@login_required
@role_required("finance", "admin", "super_admin")
def export_ventes_excel(request):
    """Export Excel des ventes."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Ventes BTA"

    entetes = ["Référence", "Client", "Formation", "Montant", "Statut", "Date"]
    ws.append(entetes)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0B2447", fill_type="solid")

    commandes = Order.objects.filter(statut="paye").prefetch_related("items")
    for cmd in commandes:
        for item in cmd.items.all():
            ws.append(
                [
                    cmd.reference,
                    cmd.utilisateur.username,
                    item.nom_produit_snapshot,
                    float(item.prix_unitaire),
                    cmd.get_statut_display(),
                    cmd.date_paiement.strftime("%d/%m/%Y") if cmd.date_paiement else "",
                ]
            )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="ventes_bta.xlsx"'
    wb.save(response)
    return response


@login_required
@role_required("finance", "admin", "super_admin")
def export_ventes_pdf(request):
    """Export PDF des ventes."""
    try:
        from weasyprint import HTML
    except (ImportError, OSError):
        messages.warning(request, "📄 L'export PDF n'est pas disponible. Utilisez l'export Excel.")
        return redirect("/admin/dashboard-business/")

    commandes = Order.objects.filter(statut="paye").prefetch_related("items")
    ca_total = commandes.aggregate(total=Sum("total"))["total"] or 0

    html_string = render_to_string(
        "academie/pdf/rapport_ventes.html",
        {
            "commandes": commandes,
            "ca_total": ca_total,
            "date_generation": timezone.now(),
        },
    )
    pdf = HTML(string=html_string).write_pdf()
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="rapport_ventes_bta.pdf"'
    return response


# ================================================
# Gradebook — Suivi des étudiants par formation
# ================================================

@login_required
@role_required("formateur", "resp_academique", "admin", "super_admin")
def gradebook_formation(request, formation_id):
    formation = get_object_or_404(Formation, id=formation_id)

    etudiants = User.objects.filter(
        acces_debloques__formation=formation
    ).distinct()

    data = []
    for etudiant in etudiants:
        note_entry = GradebookEntry.objects.filter(
            formation=formation, etudiant=etudiant
        ).first()
        progression = formation.progression_pour(etudiant)
        data.append({
            'etudiant': etudiant,
            'note': note_entry.note if note_entry else None,
            'appreciation': note_entry.appreciation if note_entry else '',
            'progression': progression,
        })

    return render(request, 'academie/gradebook.html', {
        'formation': formation,
        'etudiants': data,
        'title': f'Gradebook — {formation.nom}',
    })


@login_required
@role_required("formateur", "resp_academique", "admin", "super_admin")
def gradebook_edit(request, formation_id, etudiant_id):
    """Formulaire d'édition d'une note pour un étudiant."""
    if request.method == 'POST':
        note = request.POST.get('note')
        appreciation = request.POST.get('appreciation', '')
        try:
            note_decimal = Decimal(note)
            if note_decimal < 0 or note_decimal > 20:
                messages.error(request, "La note doit être entre 0 et 20.")
                return redirect('gradebook_formation', formation_id=formation_id)
        except:
            messages.error(request, "Note invalide.")
            return redirect('gradebook_formation', formation_id=formation_id)

        GradebookEntry.objects.update_or_create(
            formation_id=formation_id,
            etudiant_id=etudiant_id,
            defaults={
                'note': note_decimal,
                'appreciation': appreciation,
                'formateur': request.user,
            }
        )
        messages.success(request, "✅ Note enregistrée.")
        return redirect('gradebook_formation', formation_id=formation_id)

    etudiant = get_object_or_404(User, id=etudiant_id)
    formation = get_object_or_404(Formation, id=formation_id)
    entry = GradebookEntry.objects.filter(formation=formation, etudiant=etudiant).first()

    return render(request, 'academie/gradebook_edit.html', {
        'formation': formation,
        'etudiant': etudiant,
        'note': entry.note if entry else '',
        'appreciation': entry.appreciation if entry else '',
    })


# ================================================
# MENTORAT — Disponibilités du formateur
# ================================================

@login_required
@role_required("formateur", "resp_academique", "admin", "super_admin")
def mentorat_disponibilites(request):
    """Liste et gestion des disponibilités du formateur connecté."""
    formateur = request.user
    disponibilites = DisponibiliteMentor.objects.filter(formateur=formateur).order_by('date', 'heure_debut')

    if request.method == 'POST':
        date = request.POST.get('date')
        heure_debut = request.POST.get('heure_debut')
        heure_fin = request.POST.get('heure_fin')
        if date and heure_debut and heure_fin:
            DisponibiliteMentor.objects.create(
                formateur=formateur,
                date=date,
                heure_debut=heure_debut,
                heure_fin=heure_fin,
                actif=True
            )
            messages.success(request, "✅ Créneau ajouté.")
        else:
            messages.error(request, "❌ Tous les champs sont obligatoires.")
        return redirect('mentorat_disponibilites')

    return render(request, 'academie/mentorat_disponibilites.html', {
        'disponibilites': disponibilites,
        'title': 'Mes disponibilités',
    })


@login_required
@role_required("formateur", "resp_academique", "admin", "super_admin")
def mentorat_disponibilite_supprimer(request, disponibilite_id):
    """Supprime une disponibilité."""
    dispo = get_object_or_404(DisponibiliteMentor, id=disponibilite_id, formateur=request.user)
    if request.method == 'POST':
        dispo.delete()
        messages.success(request, "✅ Créneau supprimé.")
    return redirect('mentorat_disponibilites')


@login_required
def mentorat_reservations(request):
    """Liste des réservations pour l'étudiant ou le formateur."""
    if request.user.profil.role in ['formateur', 'admin', 'resp_academique']:
        reservations = ReservationMentorat.objects.filter(
            disponibilite__formateur=request.user
        ).select_related('etudiant', 'disponibilite').order_by('-date_reservation')
    else:
        reservations = ReservationMentorat.objects.filter(
            etudiant=request.user
        ).select_related('disponibilite__formateur').order_by('-date_reservation')

    return render(request, 'academie/mentorat_reservations.html', {
        'reservations': reservations,
        'title': 'Mes réservations',
    })


@login_required
def mentorat_reserver(request, disponibilite_id):
    """Un étudiant réserve un créneau."""
    if request.method == 'POST':
        dispo = get_object_or_404(DisponibiliteMentor, id=disponibilite_id, actif=True)
        if ReservationMentorat.objects.filter(disponibilite=dispo, statut__in=['en_attente', 'confirmee']).exists():
            messages.error(request, "❌ Ce créneau est déjà réservé.")
            return redirect('mentorat_calendrier')

        sujet = request.POST.get('sujet', '').strip()
        if not sujet:
            messages.error(request, "❌ Veuillez indiquer un sujet.")
            return redirect('mentorat_calendrier')

        ReservationMentorat.objects.create(
            disponibilite=dispo,
            etudiant=request.user,
            sujet=sujet,
            statut='en_attente'
        )
        messages.success(request, "✅ Demande de réservation envoyée.")
        return redirect('mentorat_reservations')

    return redirect('mentorat_calendrier')


@login_required
@role_required("formateur", "resp_academique", "admin", "super_admin")
def mentorat_changer_statut(request, reservation_id):
    """Le formateur accepte ou refuse une réservation."""
    if request.method == 'POST':
        reservation = get_object_or_404(
            ReservationMentorat,
            id=reservation_id,
            disponibilite__formateur=request.user
        )
        nouveau_statut = request.POST.get('statut')
        if nouveau_statut in ['confirmee', 'annulee', 'terminee']:
            reservation.statut = nouveau_statut
            reservation.save()
            messages.success(request, f"✅ Statut mis à jour : {reservation.get_statut_display()}")
        else:
            messages.error(request, "❌ Statut invalide.")
    return redirect('mentorat_reservations')


@login_required
def mentorat_calendrier(request):
    """Calendrier des disponibilités pour les étudiants."""
    from datetime import date, timedelta

    aujourdhui = date.today()
    fin_periode = aujourdhui + timedelta(days=7)

    disponibilites = DisponibiliteMentor.objects.filter(
        date__gte=aujourdhui,
        date__lte=fin_periode,
        actif=True
    ).select_related('formateur').order_by('date', 'heure_debut')

    return render(request, 'academie/mentorat_calendrier.html', {
        'disponibilites': disponibilites,
        'title': 'Réserver un mentorat',
    })


# ================================================
# Admin / Dashboards
# ================================================

@login_required
@role_required("resp_academique", "direction", "admin", "super_admin")
def vue_dashboard_ia(request):
    from django.contrib.auth.models import User
    from django.db.models import Count, Q, Sum

    from ..models import Article, Order, ResultatQuiz
    from ..models import Formation, Lecon

    formations_stats = (
        Formation.objects.filter(actif=True)
        .annotate(
            nb_inscrits=Count("orderitem", filter=Q(orderitem__commande__statut="paye")),
        )
        .order_by("-nb_inscrits")[:10]
    )

    lecons_abandon = Lecon.objects.annotate(
        nb_vues=Count("progressions"),
        nb_terminees=Count("progressions", filter=Q(progressions__terminee=True)),
    ).filter(nb_vues__gt=0)

    quiz_difficiles = (
        ResultatQuiz.objects.values("quiz__titre")
        .annotate(score_moyen=Avg("score"), nb_tentatives=Count("id"))
        .order_by("score_moyen")[:5]
    )

    articles_populaires = Article.objects.filter(publie=True).order_by("-date_publication")[:5]

    contexte_donnees = f"""
Formations les plus vendues : {[(f.nom, f.nb_inscrits) for f in formations_stats[:5]]}
Quiz avec scores les plus bas (difficiles) : {list(quiz_difficiles)}
Nombre total de formations actives : {Formation.objects.filter(actif=True).count()}
Nombre d'étudiants inscrits : {User.objects.filter(is_staff=False).count()}
Chiffre d'affaires : {Order.objects.filter(statut="paye").aggregate(t=Sum("total"))["t"] or 0} $
"""

    analyse_ia = None
    if request.GET.get("lancer_analyse"):
        analyse_ia = analyser_plateforme_ia(contexte_donnees)
    if analyse_ia:
        analyse_ia = analyse_ia.replace("##", "").replace("**", "").replace("---", "")
    return render(
        request,
        "admin/dashboard_ia.html",
        {
            "title": "🧠 Dashboard IA — Intelligence Décisionnelle",
            "formations_stats": formations_stats,
            "quiz_difficiles": quiz_difficiles,
            "articles_populaires": articles_populaires,
            "analyse_ia": analyse_ia,
        },
    )


@login_required
@role_required("direction", "admin", "super_admin")
def statistiques(request):
    from django.db.models import Sum

    total_etudiants = User.objects.filter(is_active=True).count()
    total_formations = Formation.objects.filter(actif=True).count()
    total_certificats = Certificat.objects.count()
    total_leads = Inscription.objects.filter(formation__gratuit=True).count()
    total_inscriptions = Inscription.objects.count()
    taux_conversion = 0
    if total_leads > 0:
        taux_conversion = round((total_inscriptions / total_leads) * 100)

    telechargements = total_leads
    comptes_crees = User.objects.count()
    inscriptions_payantes = Inscription.objects.filter(formation__gratuit=False).count()
    certificats_delivres = total_certificats
    taux_leads_comptes = (
        round((comptes_crees / telechargements) * 100) if telechargements > 0 else 0
    )
    taux_comptes_payant = (
        round((inscriptions_payantes / comptes_crees) * 100) if comptes_crees > 0 else 0
    )
    taux_payant_certif = (
        round((certificats_delivres / inscriptions_payantes) * 100)
        if inscriptions_payantes > 0
        else 0
    )

    total_revenus = (
        Inscription.objects.filter(formation__gratuit=False).aggregate(
            total=Sum("formation__prix")
        )["total"]
        or 0
    )

    debut_mois = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    certificats_mois = Certificat.objects.filter(date_emission__gte=debut_mois).count()

    quiz_generees = Quiz.objects.count()
    contenus_generes = Lecon.objects.filter(contenu__isnull=False).exclude(contenu="").count()
    formations_populaires = (
        Formation.objects.filter(actif=True)
        .annotate(nb_inscriptions=Count("inscriptions"))
        .order_by("-nb_inscriptions")[:5]
    )
    douze_mois = timezone.now() - timedelta(days=365)
    inscriptions_mensuelles = (
        Inscription.objects.filter(date_inscription__gte=douze_mois)
        .annotate(mois=TruncMonth("date_inscription"))
        .values("mois")
        .annotate(total=Count("id"))
        .order_by("mois")
    )
    mois_labels = [item["mois"].strftime("%b %Y") for item in inscriptions_mensuelles]
    mois_data = [item["total"] for item in inscriptions_mensuelles]

    total_sujets = Sujet.objects.count()
    total_reponses = Reponse.objects.count()
    membres_actifs = User.objects.filter(
        Q(id__in=Sujet.objects.values('auteur')) | Q(id__in=Reponse.objects.values('auteur'))
    ).distinct().count()

    alertes = []
    inscriptions_non_traitees = Inscription.objects.filter(traite=False).count()
    if inscriptions_non_traitees > 0:
        alertes.append(f"{inscriptions_non_traitees} inscriptions non traitées")
    formations_sans_modules = Formation.objects.filter(actif=True, modules__isnull=True).count()
    if formations_sans_modules > 0:
        alertes.append(f"{formations_sans_modules} formations sans modules")
    etudiants_inactifs = User.objects.filter(is_active=True, progressions__isnull=True).count()
    if etudiants_inactifs > 0:
        alertes.append(f"{etudiants_inactifs} étudiants inactifs")

    contexte = {
        "total_etudiants": total_etudiants,
        "total_formations": total_formations,
        "total_certificats": total_certificats,
        "total_revenus": total_revenus,
        "total_leads": total_leads,
        "total_inscriptions": total_inscriptions,
        "taux_conversion": taux_conversion,
        "formations_populaires": formations_populaires,
        "mois_labels": mois_labels,
        "mois_data": mois_data,
        "total_sujets": total_sujets,
        "total_reponses": total_reponses,
        "membres_actifs": membres_actifs,
        "alertes": alertes,
        "telechargements": telechargements,
        "comptes_crees": comptes_crees,
        "inscriptions_payantes": inscriptions_payantes,
        "certificats_delivres": certificats_delivres,
        "taux_leads_comptes": taux_leads_comptes,
        "taux_comptes_payant": taux_comptes_payant,
        "taux_payant_certif": taux_payant_certif,
        "certificats_mois": certificats_mois,
        "quiz_generees": quiz_generees,
        "contenus_generes": contenus_generes,
    }
    return render(request, "academie/statistiques.html", contexte)


@login_required
@role_required("finance", "direction", "admin", "super_admin")
def vue_dashboard_business(request):
    import json
    from datetime import timedelta

    total_inscriptions = Inscription.objects.count()
    inscriptions_non_traitees = Inscription.objects.filter(traite=False).count()

    ca_total = Order.objects.filter(statut="paye").aggregate(total=Sum("total"))["total"] or 0

    labels_jours, valeurs_jours = [], []
    for i in range(29, -1, -1):
        jour = timezone.now().date() - timedelta(days=i)
        montant_jour = (
            Order.objects.filter(statut="paye", date_paiement__date=jour).aggregate(
                total=Sum("total")
            )["total"]
            or 0
        )
        labels_jours.append(jour.strftime("%d/%m"))
        valeurs_jours.append(float(montant_jour))

    formations_populaires = Formation.objects.annotate(
        nb_ventes=Count("orderitem", filter=Q(orderitem__commande__statut="paye"))
    ).order_by("-nb_ventes")[:8]

    repartition_moyens = (
        Transaction.objects.filter(statut="reussie")
        .values("moyen_paiement__nom_affiche")
        .annotate(total=Count("id"))
    )

    coupons_utilises = Coupon.objects.filter(utilisations_actuelles__gt=0).count()
    remboursements_total = (
        Refund.objects.filter(statut="approuve").aggregate(total=Sum("montant"))["total"] or 0
    )

    return render(
        request,
        "admin/dashboard_business.html",
        {
            "title": "💼 Dashboard Business",
            "site_header": admin.site.site_header,
            "ca_total": ca_total,
            "total_inscriptions": total_inscriptions,
            "inscriptions_non_traitees": inscriptions_non_traitees,
            "formations_populaires": formations_populaires,
            "coupons_utilises": coupons_utilises,
            "remboursements_total": remboursements_total,
            "chart_labels_json": json.dumps(labels_jours),
            "chart_valeurs_json": json.dumps(valeurs_jours),
            "chart_moyens_labels": json.dumps(
                [m["moyen_paiement__nom_affiche"] or "N/A" for m in repartition_moyens]
            ),
            "chart_moyens_valeurs": json.dumps([m["total"] for m in repartition_moyens]),
        },
    )


@login_required
@role_required("formateur", "resp_academique", "admin", "super_admin")
def dashboard_enseignant(request):
    try:
        enseignant = request.user.profil.enseignant
    except (AttributeError, Enseignant.DoesNotExist):
        messages.error(request, "❌ Aucun profil enseignant associé à ce compte.")
        return redirect("dashboard")

    formations = enseignant.formations_attribuees.all()

    return render(
        request,
        "academie/dashboard_enseignant.html",
        {
            "enseignant": enseignant,
            "formations": formations,
            "revenus": enseignant.revenus_generes(),
            "remuneration": enseignant.part_remuneration(),
            "nb_etudiants": enseignant.nb_etudiants_formes(),
            "note_moyenne": enseignant.note_moyenne_temoignages(),
        },
    )


@extend_schema(
    tags=["Back Office"],
    description="Assistant IA pour les administrateurs – réponse en HTML",
    request=inline_serializer(
        name="AssistantBackOfficeRequest", fields={"question": serializers.CharField()}
    ),
    responses={
        200: inline_serializer(
            name="AssistantBackOfficeResponse", fields={"reponse": serializers.CharField()}
        )
    },
)
@api_view(["POST"])
@permission_classes([IsAuthenticated])
@login_required
@role_required(
    "formateur",
    "resp_academique",
    "marketing",
    "support",
    "finance",
    "direction",
    "admin",
    "super_admin",
)
def api_assistant_backoffice(request):
    question = request.data.get("question", "").strip()
    if not question:
        return Response({"erreur": "Question vide"}, status=400)

    contexte = (
        f"Rôle: {request.user.profil.get_role_display()}, Utilisateur: {request.user.username}"
    )

    reponse_brute = assistant_backoffice_ia(question, contexte)

    try:
        reponse_html = markdown_lib.markdown(reponse_brute)
    except Exception:
        reponse_html = reponse_brute

    return Response({"reponse": reponse_html})